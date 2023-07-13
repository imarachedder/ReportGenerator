from openpyxl import load_workbook
import xlwings as xw
import tkinter as tk
import tkinter.filedialog as fd
from datetime import date
import db

db = db.Query()
list_obj = db.get_road_name()
db.db_close()
columns = list(list_obj.keys())
values = list(list_obj.values())
# for i in range(3):
#    print(len(values[i]))
fn = r'C:\Users\sibregion\Desktop\мм\Новая папка\test.xlsx'
wb = load_workbook(fn, keep_vba = True)
sheet_names = wb.sheetnames
if 'DB' in sheet_names:
    ws = wb['DB']
else:
    ws = wb.create_sheet(title = 'DB', index = None)
# print(len(columns), len(values[0]))
ws.append(columns)
for j in range(len(columns)):
    for i in range(len(values[0])):
        ws.cell(row = i + 2, column = j + 1).value = values[j][i]
        # print(values[j][i])
wb.save(r'C:\Users\sibregion\Desktop\мм\Новая папка\newfile.xlsm')
wb.close()
