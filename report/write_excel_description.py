import os

import xlsxwriter
from openpyxl import load_workbook
import xlwings as xw
from datetime import date
import db
from interface import window2, window3


class Report:
    def __init__ (self):
        self.db = db.Query()
        self.path_template = r"C:\Users\sibregion\Desktop\test\report\Новая папка\ТП2.xlsx"
        self.info_window2 = window2.Window2().get_info_window2()
        self.info_window3 = window3.Window3().get_info_from_plainTextEdit()
        self.wb = load_workbook(self.path_template, keep_vba = True)
        self.path = os.path.dirname(os.path.abspath(__file__))


    def get_info_window2 (self):
        return self.info_window2

    def write_titular (self, data_dict):
        #columns = list(data_dict.keys())
        #values = list(data_dict.values())

        # if 'DB' in sheet_names:
        #    ws = wb['DB']
        # else:
        #    ws = wb.create_sheet(title = 'DB', index = None)
        # print(len(columns), len(values[0]))
        ws = self.wb['Титульник (с рамкой)']
        # ws.append(columns)
        ws["B4"].value = data_dict['client']
        ws["B22"].value = data_dict['name_road']
        ws["B31"].value = f"составлена на {data_dict['year']} г."
        ws["B33"].value = f"Шифр:{data_dict['cypher']}"
        ws["B52"].value = f"Омск:{data_dict['year']}"
        ws["B41"].value = data_dict['contractor']
        ws["B46"].value = f'{data_dict["position_contractor"]} {data_dict["fio_contractor"]}________________________'
        ws["AI41"].value = data_dict['client']
        ws["AI46"].value = f'{data_dict["position_client"]} {data_dict["fio_client"]}________________________'

    def write_6 (self, data_dict):
        ws = self.wb['6']
        ws["B4"].value = data_dict['name_road']
        for i in range(len(data_dict['количество_участков'])):
            ws["L1" + str(i)].value = data_dict['участки'][i]
        ws["S15"].value = 'знач!'

        ws["AL10"].value = data_dict['name_road']




    def save_file (self):
        self.wb.save(r'C:\Users\sibregion\Desktop\test\report\Новая папка\ТП2.xlsx')
    def close_file (self):
        self.wb.close()
def write_excel_openpyxl (fn = r'C:\Users\sibregion\Desktop\мм\Новая папка\test.xlsm', data = None):
    list_obj = db.get_road_name()
    histori_match = 'histori_match'
    railway_waterway = 'railway_waterway'
    economical_characteristic_road = 'economical_characteristic_road'
    movement_characteristic = 'movement_characteristic'
    area_conditions = 'area_conditions'
    client = 'client'
    fio_client = 'fio_client'
    position_client = 'position_client'
    contractor = 'contractor'
    fio_contractor = 'fio_contractor'
    position_contractor = 'position_contractor'
    year = 'year'
    cypher = 'cypher'

    data_dict = {'client': client,
                 'fio_client': fio_client,
                 'position_client': position_client,
                 'contractor': contractor,
                 'fio_contractor': fio_contractor,
                 'position_contractor': position_contractor,
                 'year': year,
                 'cypher': cypher,
                 'histori_match': histori_match,
                 'railway_waterway': railway_waterway,
                 'economical_characteristic_road': economical_characteristic_road,
                 'movement_characteristic': movement_characteristic,
                 'area_conditions': area_conditions}

    columns = list(list_obj.keys())
    values = list(list_obj.values())
    # for i in range(3):
    #    print(len(values[i]))
    fn = r'C:\Users\sibregion\Desktop\test\report\Новая папка\test.xlsm'
    wb = load_workbook(fn, keep_vba = True)
    sheet_names = wb.sheetnames
    if 'DB' in sheet_names:
        ws = wb['DB']
    else:
        ws = wb.create_sheet(title = 'DB', index = None)
    # print(len(columns), len(values[0]))
    ws.append(columns)
    for j in range(len(columns)):
        for i in range(len(values[0])):
            ws.cell(row = i + 2, column = j + 1).value = values[j][i]
            # print(values[j][i])
    wb.save(os.path.dirname(os.path.abspath(__file__)) + r'\newfile.xlsm')
    wb.close()


histori_match = 'histori_match'
railway_waterway = 'railway_waterway'
economical_characteristic_road = 'economical_characteristic_road'
movement_characteristic = 'movement_characteristic'
area_conditions = 'area_conditions'
client = 'client'
fio_client = 'fio_client'
position_client = 'position_client'
contractor = 'contractor'
fio_contractor = 'fio_contractor'
position_contractor = 'position_contractor'
year = 'year'
cypher = 'cypher'



def write_excel_xlwings_titular (filename):
    wb = xw.Book(filename)
    ws = wb.sheets['Титульник (с рамкой)']
    ws.range("B22").value = data_dict['name_road']
    ws.range("B31").value = f"составлена на {data_dict['year']} г."
    ws.range("B33").value = f"Шифр:{data_dict['cypher']}"
    ws.range("B52").value = f"Омск:{data_dict['year']}"
    ws.range("B41").value = data_dict['contractor']
    ws.range("B46").value = f'{data_dict["position_contractor"]} {data_dict["fio_contractor"]}________________________'
    ws.range("AI41").value = data_dict['client']
    ws.range("AI46").value = f'{data_dict["position_client"]} {data_dict["fio_client"]}________________________'
    wb.save(filename)

data = {
        'name_road': 'name_road',
        'client': client,
        'fio_client': fio_client,
        'position_client': position_client,
        'contractor': contractor,
        'fio_contractor': fio_contractor,
        'position_contractor': position_contractor,
        'year': year,
        'cypher': cypher,
        'histori_match': histori_match,
        'railway_waterway': railway_waterway,
        'economical_characteristic_road': economical_characteristic_road,
        'movement_characteristic': movement_characteristic,
        'area_conditions': area_conditions}



if __name__ == "__main__":

    Report.write_titular(data)
    #write_excel_xlwings(r"C:\Users\sibregion\Desktop\test\report\Новая папка\ТП2.xlsx")
    # db = db.Query()
    # list_obj = db.get_road_name()
    # db.db_close()
    # columns = list(list_obj.keys())
    # values = list_obj.values()

    # print(data)
    # write_excel_openpyxl(data = data)
    # write_excel_openpyxl(data = values)
    # write_excel_openpyxl(data = list_obj)

    # write_excel_openpyxl(data = list_obj)
    # write_excel_xlwings(r'C:\Users\sibregion\Desktop\мм\Новая папка\test.xlsx')
    # app = App()
    # app.mainloop()

# structure = {'id':,'measurement':{'name_measurement':1 и тд} }
'''
    wb = load_workbook(fn, keep_vba = True)
    list_sheets = wb.sheetnames
    print(list_sheets)
    #['Данные', 'Описание', 'Титульник (без рамки)', 'Титульник (с рамкой)', 'Содержание', 'Схема', '6', '7', '8', '9', '10', '11', '12', '12 (4)', '13', '13 (2)', '13 (3)', '14', '15', '16', '17', '18', '19']
    ws = wb.create_sheet(title='DB', index=None)
    #ws["E6"] = "Какая то дорога186y77"
    #ws.cell(row=1, column=10).value = date.today().year
    if data is not None:
        ws.append(data)
    wb.save(r'C:\\Users\sibregion\Desktop\мм\Новая папка\\newfile.xlsm')
    wb.close()
'''
