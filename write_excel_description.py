import math
from pathlib import Path

import win32com.client
from PyQt6.QtWidgets import QMessageBox
from icecream import ic
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.text import RichText
from openpyxl.drawing.image import Image
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties, Paragraph, Font as FontText, \
    RichTextProperties
from openpyxl.styles import Alignment, Side, Border, Font as FontStyle

import db
from settings import path_template_excel, path_template_excel_application, path_template_excel_dad


class WriterExcel:
    def __init__ (self, data: dict = None, path_template_excel = path_template_excel, path = None,
                  data_interface = None):

        self.tip_doc = None
        self.data = data
        if data is None:
            self.data = {}
        self.wb = load_workbook(filename = path_template_excel, keep_vba = True, keep_links = True)
        self.path_dir = Path(path)
        self.errors = open(self.path_dir.joinpath('Ошибки.txt'), "a", encoding = "utf-8")
        self.data_interface = data_interface
        if data_interface is None:
            self.data_interface = {'tip_passport': 'city'}
        self.table_cells_font = FontStyle(name = 'Times New Roman', size = 12)
        thin = Side(border_style = "thin", color = "000000")
        self.table_cells_border = Border(left = thin, right = thin, top = thin, bottom = thin, )
        self.table_cells_aligment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)

    def save_file (self):
        # сохранить файл
        name_file = self.data.get('название дороги', 'дорога')
        if '/' in name_file or ':' in name_file:
            name_file = name_file.replace("/", ".").replace(":", ".")[:40]
        try:
            self.wb.save(rf'{self.path_dir}\{self.tip_doc}_{name_file}.xlsm')
        except Exception as e:
            QMessageBox.information(None, f"Ошибка", f'Ошибка сохранения файла: {e}',
                                    QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.NoButton)
        self.close_file()

    def close_file (self):
        # закрыть файл
        self.wb.close()

    def change_start_and_end_obj (self, start_obj, end_obj):
        '''
        Преобразует начало и конец объекта в строковый вид  км+м(000)
        :param start_obj:
        :param end_obj:
        :return: возвращает начало и конец в строковом виде км+м(000)
        '''
        if 9 < start_obj[1] < 100:
            str_start_obj = f'{start_obj[0]}+0{start_obj[1]}'
        elif start_obj[1] < 10:
            str_start_obj = f'{start_obj[0]}+00{start_obj[1]}'
        else:
            str_start_obj = f'{start_obj[0]}+{start_obj[1]}'
        if 9 < end_obj[1] < 100:
            str_end_obj = f'{end_obj[0]}+0{end_obj[1]}'
        elif end_obj[1] < 10:
            str_end_obj = f'{end_obj[0]}+00{end_obj[1]}'
        else:
            str_end_obj = f'{end_obj[0]}+{end_obj[1]}'
        return str_start_obj, str_end_obj


class WriterExcelTP(WriterExcel):
    def __init__ (self, data: dict = None, path = None, data_interface = None):
        super().__init__(data = data, path = path, data_interface = data_interface)
        self.tip_doc = 'ТП'
        self.run()
        self.errors.close()
        print("сохранение")
        self.save_file()

    def run (self):
        print("Начал работать")
        # print(self.data_interface)

        # # self.write_pereplet()
        print('титульный')
        self.write_titular()
        print('схема')
        self.write_scheme()
        print('6 лист')
        self.write_6()
        print('7 лист')
        self.write_7()
        print('8 лист')
        self.write_8()
        print('9 лист')
        self.write_9()
        print('10 лист')
        self.write_10()
        print('11 лист')
        self.write_11()
        print('12 лист')
        self.write_12()
        print('13 лист')
        self.write_13()
        print('14 лист')
        self.write_14()
        print('17 лист')
        self.write_17()
        print('18 лист')
        self.write_18()

    def write_pereplet (self):
        """
        Заполняет Титульник для переплета
        :return:
        """
        ws = self.wb['Переплет']  # выбираем лист
        # ws.header_footer
        ws['B22'].value = self.data.get('название дороги')

    def write_titular (self):
        """
        Заполняет лист 'Титульник (без рамки)'
        :return:
        """
        ws = self.wb['Титульник (без рамки)']  # выбираем лист

        ws["B4"].value = self.data_interface.get('client', 'Заказчик')
        ws["B22"].value = self.data.get('название дороги', 'дорога')
        ws["B31"].value = f"составлена на {self.data_interface.get('year', 'год')} г."
        ws["B33"].value = f"Шифр:{self.data_interface.get('cypher', 'шифр')} "
        ws["B52"].value = f"Омск - {self.data_interface.get('year', 'год')} г."
        ws["B41"].value = self.data_interface.get('contractor', 'Подрядчик')
        ws["B46"].value = f'{self.data_interface.get("position_contractor", "Должность подрядчика")} ' \
                          f'{self.data_interface.get("fio_contractor", "ФИО подрядчика")}________________________'
        ws["AI41"].value = self.data_interface.get('client', 'Заказчик')
        ws[
            "AI46"].value = f'{self.data_interface.get("position_client", "Должность заказчика")} {self.data_interface.get("fio_client", "ФИО заказчика")}' \
                            f'________________________'

    def write_scheme (self):
        """
         Заполняет лист "схема"
        :return: None
        """
        ws = self.wb['Схема']  # выбираем лист

        try:
            schema = Image(f"{self.path_dir}\Схема.png")
            # schema.width = 1380
            # schema.height = 800
            ws.add_image(schema, 'B5')

        except FileNotFoundError:
            self.errors.write(f'{self.tip_doc} Cхема: Файл Схема.png отсутствует в {self.path_dir}\n')
        # self.msg.information(self.parent, "Файл не найден  ",
        #                      f'Файл Схема.png отсутствует в {self.path_dir} \n{e}',  QMessageBox.StandardButton.Ok,
        #                      QMessageBox.StandardButton.NoButton)

    def write_6 (self):
        """
        Заполняет лист "6"
        :return:
        """

        ws = self.wb['6']  # выбираем лист
        n, i = 9, 21  # счетчик
        j = 10
        res = 0
        # 2.1 Наименование дороги: name road
        ws["O5"].value = self.data.get('название дороги', 'дорога')

        # 2.2 Участок дороги 1, 2 и т.д., 2.3 протяженность дороги(участка) и 2.5 категория дороги(участка), подъездов
        ws["AL10"] = self.data.get('название дороги', 'дорога')

        for key, value in self.data.items():
            if isinstance(value, str):
                continue
            id_key = list(self.data.keys()).index(key)
            try:
                start_road = value.get('Ось дороги', {}).get('Начало трассы', [])[0][-2]
                end_road = value.get('Ось дороги', {}).get('Начало трассы', [])[0][-1]
            except Exception:
                self.errors.write(f'{self.tip_doc} 6 лист: Проверьте объект "Ось дороги" на участке {key}\n')
                # self.msg.information(self.parent, f"{self.tip_doc} 6 лист",
                #                      f'Проверьте объект "Ось дороги" на участке {key}',
                #                      QMessageBox.StandardButton.Ok,
                #                      QMessageBox.StandardButton.NoButton)

            str_start, str_end = self.change_start_and_end_obj(start_road, end_road)
            ws[f'B{n}'] = f'2.2 Участок дороги:'
            if len(self.data) > 2:
                ws[f'B{n}'] = f'2.2 Участок дороги {id_key}:' if n == 9 else f'Участок дороги {id_key}:'
                ws.unmerge_cells(start_row = i, start_column = 2, end_row = i, end_column = 5)
                ws.unmerge_cells(start_row = i, start_column = 6, end_row = i, end_column = 9)
                ws.unmerge_cells(start_row = i, start_column = 10, end_row = i, end_column = 13)
                ws.unmerge_cells(start_row = i, start_column = 14, end_row = i, end_column = 17)
                ws.unmerge_cells(start_row = i, start_column = 18, end_row = i, end_column = 21)
                ws.unmerge_cells(start_row = i, start_column = 22, end_row = i, end_column = 27)
                ws.unmerge_cells(start_row = i, start_column = 28, end_row = i, end_column = 32)
                ws.unmerge_cells(start_row = i, start_column = 33, end_row = i, end_column = 36)
                ws.merge_cells(f'B{i}:AJ{i}')
                for col in range(2, 37):
                    ws.cell(row = i, column = col).border = self.table_cells_border
                    ws.cell(row = i, column = col).alignment = self.table_cells_aligment

                ws[f'B{i}'] = key.title()
                for col in range(49, 68):
                    ws.cell(row = j, column = col).border = self.table_cells_border
                    ws.cell(row = j, column = col).alignment = self.table_cells_aligment
                ws.merge_cells(f'AW{j}:BO{j}')
                ws[f'AW{j}'].alignment = self.table_cells_aligment

                ws[f'AW{j}'] = key.title()
                i += 1
                j += 1

            ws[f'B{i}'] = f"{str_start}"
            ws[f'F{i}'] = f"{str_end}"
            length_district = round(
                (end_road[0] * 1000 + end_road[1] - int(
                    value.get('Ось дороги', {}).get('Начало трассы', [])[0][0])) / 1000,
                3)
            ws[f'J{i}'] = length_district

            ws[f'L{n}'] = f"от КМ {str_start} до КМ {str_end}"
            try:
                tuple_cateregory = value.get('Граница участка дороги', {}).get('категория а/д')
            except KeyError as e:
                tuple_cateregory = []
                self.errors.write(
                    f'{self.tip_doc} 6 лист: Проверьте данные в объекте "Граница участка дороги" в поле категория А/Д\n')

            last_cat = None
            for idx, category in enumerate(tuple_cateregory):
                if category[-1] == end_road:
                    # если конец дороги
                    ws[f'AW{j - 2}'].alignment = self.table_cells_aligment
                    ws[f'AW{j - 2}'].border = self.table_cells_border
                    ws[f'AW{j - 2}'] = self.change_start_and_end_obj(last_cat[-2], last_cat[-1])[1]

                    ws[f'BA{j - 2}'].alignment = self.table_cells_aligment
                    ws[f'BA{j - 2}'].border = self.table_cells_border
                    ws[f'BA{j - 2}'] = str_end

                    ws[f'BE{j - 2}'].alignment = self.table_cells_aligment
                    ws[f'BE{j - 2}'].border = self.table_cells_border
                    ws[f'BE{j - 2}'] = category[0]
                    break
                for col in range(49, 68):
                    ws.cell(row = j, column = col).border = self.table_cells_border
                    ws.cell(row = j, column = col).alignment = self.table_cells_aligment
                ws.merge_cells(f'AW{j}:AZ{j + 1}')
                ws.merge_cells(f'BA{j}:BD{j + 1}')
                ws.merge_cells(f'BE{j}:BO{j + 1}')

                ws[f'AW{j}'].alignment = self.table_cells_aligment
                ws[f'AW{j}'].border = self.table_cells_border
                ws[f'BA{j}'].alignment = self.table_cells_aligment
                ws[f'BA{j}'].border = self.table_cells_border
                ws[f'BE{j}'].alignment = self.table_cells_aligment
                ws[f'BE{j}'].border = self.table_cells_border
                if (len(tuple_cateregory) == 2 and tuple_cateregory[0][0] == tuple_cateregory[-1][0]):
                    # если 2 одинаковых категории

                    ws[f'AW{j}'] = str_start
                    ws[f'BA{j}'] = str_end
                    ws[f'BE{j}'] = category[0]
                    j += 2
                    break
                if idx == 0:
                    # если первый элемент
                    ws[f'AW{j}'] = str_start
                    ws[f'BA{j}'] = self.change_start_and_end_obj(category[-2], category[-1])[1]
                    ws[f'BE{j}'] = category[0]
                    last_cat = category

                elif category[0] == last_cat[0]:
                    # если категори совпадает с пердыдущей
                    ws[f'BA{j - 2}'] = self.change_start_and_end_obj(category[-2], category[-1])[1]
                    last_cat = category

                else:
                    ws[f'AW{j}'] = self.change_start_and_end_obj(category[-2], category[-1])[0]
                    ws[f'BA{j - 2}'] = self.change_start_and_end_obj(category[-2], category[-1])[1]
                    ws[f'BE{j}'] = category[0]
                    last_cat = category
                j += 2

            res += length_district
            n += 1
            i += 1
        res = round(res, 3)
        ws['S14'] = f"{res} км"

        # except Exception as e:
        #     print('Ошибка заполнения 6 листа', e)
        #     self.save_file()
        #     raise e

        # self.msg.setText("Ошибка заполнения 6 листа. Обратитесь к разработчику!")
        # self.msg.setWindowTitle("Ошибка - 6 лист")
        # self.msg.exec()

        # заполняет таблицу 2.4 Наименование подъездов (обходов) и их протяженность
        # ws["B37"].value = self.data.get('подъезды', {}).get('Наименование', [])

        # заполняет таблицу 2.6 Краткая историческая справка
        ws["AL33"].value = self.data_interface.get('history_match', None)

    def write_7 (self):
        # 2.7
        ws = self.wb['7']
        counter_distr_soder = 15  # счетчик строк для 2.7
        column_tuple = ('AX', 'AZ', 'BB', 'BD', 'BF', 'BH', 'BJ', 'BL')  # столбцы для 2.8

        row_name_distr = 15  # счетчик строк для 2.8
        for k1, v1 in self.data.items():
            if isinstance(v1, str):
                continue
            # try:
            for idx, v2 in enumerate(v1.get('Дорожная организация', {}).get('Наименование', [])):
                ws[f'B{counter_distr_soder}'] = self.data_interface.get('year', '-')
                ws[f'E{counter_distr_soder}'] = v1.get('Дорожная организация', {}).get('Наименование', [])[idx][0] \
                    if v1.get('Дорожная организация', {}).get('Наименование', False) else '-'
                ws[f'l{counter_distr_soder}'] = v1.get('Дорожная организация', {}).get('Адрес', [])[idx][0] if \
                    v1.get('Дорожная организация', {}).get('Адрес', False) else '-'
                ws[f'P{counter_distr_soder}'] = v1.get('Дорожная организация', {}).get('Город', [])[idx][
                    0] if v1.get('Дорожная организация', {}).get('Город', False) else '-'
                ws[f'V{counter_distr_soder}'] = v1.get('Дорожная организация', {}).get('Начало по оси', [])[idx][
                    0] if v1.get('Дорожная организация', {}).get('Начало по оси', False) else '-'
                ws[f'Y{counter_distr_soder}'] = v1.get('Дорожная организация', {}).get('Конец  по оси', [])[idx][0] \
                    if v1.get('Дорожная организация', {}).get('Конец  по оси', False) else '-'
                # начало и конец по оси должны быть записаны с километровой привязкой км+м
                try:
                    start = v1.get('Дорожная организация', {}).get('Начало по оси', [])[idx][0].split('+')
                    end = v1.get('Дорожная организация', {}).get('Конец  по оси', [])[idx][0].split('+')
                except Exception:
                    start = 0
                    end = 0
                    self.errors.write(
                        f"{self.tip_doc} лист 7 таблицы 2.7: Проверьте объект 'Дорожная организация'{v1.get('Дорожная организация', {}).get('Начало по оси', )[idx][0]}")

                ws[
                    f'AB{counter_distr_soder}'] = f'{((int(end[0]) - int(start[0])) * 1000 + int(end[1]) - int(start[1])) / 1000}'
                ws[f'AK{counter_distr_soder}'] = f'=AB{counter_distr_soder}'
                counter_distr_soder += 1

            # 2.8 Таблица основных расстояний (в целых километрах)
            tuple_name = tuple(v1.get('Населенный пункт', {}).get('Наименование', []))
            try:
                for idx, name in enumerate(tuple_name):
                    ws[f'{column_tuple[idx]}4'] = name[0]
                    ws[f'AR{row_name_distr}'] = name[0]
                    # итератор столбцов, споймает ошибку если населенных пунктов будет больше чем указанных столобцов в Excel
                    iter_column = iter(column_tuple[:len(tuple_name)])
                    for name1 in tuple_name:
                        '''
                        заполнение расстояний между населенными пунктами, в целых километрах. next(iter) возвращает каждый 
                        раз новый столбец
                        '''
                        ws[f'{next(iter(iter_column))}{row_name_distr}'] = abs(
                            (int(name1[-4]) - int(name[-4])) * 1000 + int(name1[-3]) - int(name[-3])) // 1000 \
                            if name1[2] - name[2] != 0 else '-'
                    row_name_distr += 1
            except Exception:
                self.errors.write(
                    f"{self.tip_doc} лист 7 таблицf 2.8: Проверьте объект 'Населенный пункт', если их больше 8 значит таблица переполняется\n")
            #     print('Ошибка заполнения лист 7 таблицы 2.8', e)
            # self.msg.setText("Ошибка заполнения таблицы 2.8")
            # self.msg.setWindowTitle("Ошибка в листе 7")
            # self.msg.exec()

    def write_8 (self):
        """
        Расписываем экономическую характеристику
        :param data:
        :return:
        """

        # Счетчик
        n = 1

        # Выбираем лист
        ws = self.wb['8']
        # 3.1 Экономическое и административное значение дороги
        ws['B6'] = self.data_interface.get('economical_characteristic_road', '')
        # 3.2 Связь дороги с ж/д и водными путями и автомобильными дорогами
        ws['B19'] = self.data_interface.get('railway_waterway', '')
        # 3.3 Характеристика движения, его сезонность и перспектива роста
        ws['B33'] = self.data_interface.get('movement_characteristic', '')
        # 3.4 Среднесуточная интенсивность движения по данным учета

    def write_9 (self):
        """
        Техническая характеристика
        :param res: значение макс значения оси
        :param data:
        :return:
        """

        # Функция для расчета ширины проезжей части
        def calcLengthOfTheWidthOfTheCarriageWay (res, j, key, v):
            result = 0
            if v == 'Ширина земляного полотна':
                result = res - int(self.data.get(key).get(v, {}).get('Ширина')[j - 1][8])
                # if j - 1 == 0:
                #     result += int(self.data.get(key).get(v, {}).get('Ширина')[j - 1][8])
                return result

            elif v == 'Ширина проезжей части':

                if j - 1 == 0:
                    result += int(self.data.get(key).get(v, {}).get('Ширина ПЧ')[j - 1][8])
                else:
                    result = res - int(self.data.get(key).get(v, {}).get('Ширина ПЧ')[j - 1][8])
                return result

        # Счетчик
        n, i = 12, 34

        ws = self.wb['9']
        # 4.1 Топографические условия района проложения автомобильной дороги
        ws['B7'] = self.data_interface.get('area_conditions')
        # 4.2 Ширина земляного полотна
        # 4.3 Характеристика проезжей части
        # 4.3.1 Ширина проезжей части
        for key, val in self.data.items():
            if isinstance(val, str):
                continue
            else:
                if len(self.data) > 2:
                    ws[f'AJ{n}'] = key.title()
                    ws[f'B{i}'] = key.title()
                else:
                    ws[f'AJ{n}'] = f'Весь участок'
                    ws[f'B{i}'] = f'Весь участок'
                i += 1
                n += 1
                # Создаем переменные для ячеек в таблице 4.3.1 Ширина проезжей части
                res2, res3, res4, res5, res6, res7, res8, res9, res10, res11 = 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
                sum1, sum2, sum3, sum4, sum5, sum6 = 0, 0, 0, 0, 0, 0

                res = val.get('Ось дороги').get('Начало трассы')[0][8]

                # 4.2 Ширина земляного полотна
                try:
                    for j in range(len(val.get('Ширина земляного полотна', {}).get('Ширина', [])), 0, -1):
                        if float(val.get('Ширина земляного полотна', {}).get('Ширина', [])[j - 1][0]) <= 8.0:
                            sum1 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина земляного полотна')
                        elif 8.0 < float(val.get('Ширина земляного полотна', {}).get('Ширина', [])[j - 1][0]) < 10.0:
                            sum2 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина земляного полотна')
                        elif 10.0 <= float(val.get('Ширина земляного полотна', {}).get('Ширина', [])[j - 1][0]) < 12.0:
                            sum3 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина земляного полотна')
                        elif 12.0 <= float(val.get('Ширина земляного полотна', {}).get('Ширина', [])[j - 1][0]) < 15.0:
                            sum4 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина земляного полотна')
                        elif 15.0 <= float(val.get('Ширина земляного полотна', {}).get('Ширина', [])[j - 1][0]) < 27.5:
                            sum5 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина земляного полотна')
                        elif 27.5 <= float(val.get('Ширина земляного полотна', {}).get('Ширина', [])[j - 1][0]):
                            sum6 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина земляного полотна')
                        res = val.get('Ширина земляного полотна', {}).get('Ширина', [])[j - 1][8]
                        ws[f'G{i}'].value = '-' if sum1 == 0 else round(sum1 / 1000, 3)
                        ws[f'K{i}'].value = '-' if sum2 == 0 else round(sum2 / 1000, 3)
                        ws[f'P{i}'].value = '-' if sum3 == 0 else round(sum3 / 1000, 3)
                        ws[f'U{i}'].value = '-' if sum4 == 0 else round(sum4 / 1000, 3)
                        ws[f'Z{i}'].value = '-' if sum5 == 0 else round(sum5 / 1000, 3)
                        ws[f'AE{i}'].value = '-' if sum6 == 0 else round(sum6 / 1000, 3)
                except Exception as e:
                    self.errors.write(f'{self.tip_doc} лист 9: 4.2 Ширина земляного полотна {e}\n')
                    # self.msg.setText(f"Ошибка 4.2")
                    # self.msg.setWindowTitle("Ошибка в 9 листе")
                    # self.msg.exec()

                    # 4.3 Характеристика проезжей части
                    # 4.3.1 Ширина проезжей части
                res = val.get('Ось дороги').get('Начало трассы')[0][8]
                for j in range(len(val.get('Ширина проезжей части', {}).get('Ширина ПЧ', [])), 0,
                               -1):
                    if float(val.get('Ширина проезжей части', {}).get('Ширина ПЧ')[j - 1][0]) <= 4.0:
                        res2 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина проезжей части')
                    elif 4.0 < float(
                            val.get('Ширина проезжей части', {}).get('Ширина ПЧ')[j - 1][0]) < 4.5:
                        res3 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина проезжей части')
                    elif 4.5 <= float(
                            val.get('Ширина проезжей части', {}).get('Ширина ПЧ')[j - 1][0]) < 6.0:
                        res4 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина проезжей части')
                    elif 6.0 <= float(
                            val.get('Ширина проезжей части', {}).get('Ширина ПЧ')[j - 1][0]) < 6.6:
                        res5 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина проезжей части')
                    elif 6.6 <= float(
                            val.get('Ширина проезжей части', {}).get('Ширина ПЧ')[j - 1][0]) < 7.0:
                        res6 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина проезжей части')
                    elif 7.0 <= float(
                            val.get('Ширина проезжей части', {}).get('Ширина ПЧ')[j - 1][0]) < 7.5:
                        res7 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина проезжей части')
                    elif 7.5 <= float(
                            val.get('Ширина проезжей части', {}).get('Ширина ПЧ')[j - 1][0]) < 9.1:
                        res8 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина проезжей части')
                    elif 9.1 <= float(
                            val.get('Ширина проезжей части', {}).get('Ширина ПЧ')[j - 1][
                                0]) < 10.0:
                        res9 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина проезжей части')
                    elif 10.0 <= float(
                            val.get('Ширина проезжей части', {}).get('Ширина ПЧ')[j - 1][
                                0]) < 15.1:
                        res10 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина проезжей части')
                    elif 15.1 <= float(
                            val.get('Ширина проезжей части', {}).get('Ширина ПЧ')[j - 1][0]):
                        res11 += calcLengthOfTheWidthOfTheCarriageWay(res, j, key, 'Ширина проезжей части')

                    res = val.get('Ширина проезжей части').get('Ширина ПЧ')[j - 1][8]
                    ws[f'AJ{n}'].value = self.data_interface.get('year', '')
                    ws[f'AL{n}'].value = '-' if res2 == 0 else round(res2 / 1000, 3)
                    ws[f'AO{n}'].value = '-' if res3 == 0 else round(res3 / 1000, 3)
                    ws[f'AR{n}'].value = '-' if res4 == 0 else round(res4 / 1000, 3)
                    ws[f'AU{n}'].value = '-' if res5 == 0 else round(res5 / 1000, 3)
                    ws[f'AX{n}'].value = '-' if res6 == 0 else round(res6 / 1000, 3)
                    ws[f'BA{n}'].value = '-' if res7 == 0 else round(res7 / 1000, 3)
                    ws[f'BD{n}'].value = '-' if res8 == 0 else round(res8 / 1000, 3)
                    ws[f'BG{n}'].value = '-' if res9 == 0 else round(res9 / 1000, 3)
                    ws[f'BJ{n}'].value = '-' if res10 == 0 else round(res10 / 1000, 3)
                    ws[f'BM{n}'].value = '-' if res11 == 0 else round(res11 / 1000, 3)
                n += 1
                i += 1

    def write_10 (self):

        def count_coating (v):

            """
            Расчет протяженностей типов покрытий. Для расчета нужны объекты - граница участков дороги
            @param: v
            @return: type_of_coating

            """
            capital = {'Цементобетонные монолитные': 0,
                       'Железобетонные монолитные': 0,
                       'Железобетонные сборные': 0,
                       'Армобетонные монолитные': 0,
                       'Армобетонные сборные': 0,
                       'Асфальтобетонные': 0,
                       'Щебеночно-мастичные': 0
                       }
            lightweight = {
                'Асфальтобетонные': 0,
                'Органоминеральные': 0,
                'Щебеночные (гравийные), обработанные вяжущим': 0,
                'асфальтобетон': 0

            }
            transition = {
                'Щебеночно-гравийно-песчаные': 0,
                'Грунт и малопрочные каменные материалы, укрепленные вяжущим': 0,
                'Грунт, укрепленный различными вяжущими и местными материалами': 0,
                'Булыжный и колотый камень (мостовые)': 0,
                'щебень/гравий': 0
            }
            lower = {
                'Грунт': 0,
            }
            type_of_coating = {
                'Капитальный': capital,
                'Облегченный': lightweight,
                'Переходный': transition,
                'Низший': lower
            }

            tuple_tip = v.get('Граница участка дороги', {}).get('тип дорожной одежды', [])
            tuple_variant = v.get('Граница участка дороги', {}).get('вид покрытия', [])

            for idx, tip in enumerate(tuple_tip):
                # находим следующий тип дорожной одежды
                if tip == tuple_tip[-1]:
                    next_tip = tuple_tip[-1]
                elif tip == tuple_tip[0]:
                    next_tip = tuple_tip[1]
                else:
                    next_tip = tuple_tip[idx % len(tuple_tip) + 1]
                try:
                    type_of_coating[tip[0]][tuple_variant[idx][0]] += ((next_tip[-2][0] - tip[-2][0]) * 1000 + (
                            next_tip[-2][1] - tip[-2][1])) / 1000
                except Exception as e:
                    # text = ',\n'.join(list(type_of_coating[tip[0]]))
                    self.errors.write(f'{self.tip_doc} 10 лист 4.3.2 В объекте "Граница участка дороги"{tip[-1]}'
                                      f' вид покрытия "{tuple_variant[idx][0]}" не совпадает'
                                      f'с заданными в программе')

            return type_of_coating

        ws = self.wb['10']
        column_tuple = ['AF', 'AL', 'AR', 'AX', 'BD', 'BJ']
        counter = 0
        sum_res = {}
        for k1, v1 in self.data.items():
            if isinstance(v1, str):
                continue
            else:
                column = column_tuple[counter]
                if len(self.data) > 2:
                    ws[f'{column}4'] = f'Участок {counter + 1} \n {self.data_interface.get("year", None)} г.'
                else:
                    ws[f'{column}4'] = f'{self.data_interface.get("year", None)}'
                result = count_coating(v1)
                for tip, dict_material in result.items():
                    for material, val in dict_material.items():
                        if tip in sum_res:
                            if material in sum_res.get(tip):
                                sum_res[tip][material] += val
                            else:
                                sum_res.get(tip).update({material: val})
                        else:
                            sum_res.update({tip: {material: val}})

                row = 8
                for key, val in result.items():
                    # заполнение всех типов и видов покрытий
                    ws[f'B{row - 1}'] = key
                    for key_material, material in val.items():
                        ws[f'B{row}'] = key_material
                        ws[f'{column}{row}'] = material if material != 0 else '-'
                        row += 1
                    row += 4
                # ws[f'{column}{row}'] = result.get('Капитальный').get('Цементобетонные монолитные') if result.get(
                #         'Капитальный').get('цементобетон') != 0 else '-'
                # ws[f'{column}8'] = result.get('Капитальный').get('Цементобетонные монолитные') if result.get(
                #     'Капитальный').get('цементобетон') != 0 else '-'
                # ws[f'{column}9'] = result.get('Капитальный').get('ж/б плиты') if result.get('Капитальный').get(
                #     'ж/б плиты') != 0 else ''
                # ws[f'{column}10'] = result.get('Капитальный').get('цементобетон') if result.get(
                #     'Капитальный').get('цементобетон') != 0 else '-'
                # ws[f'{column}11'] = result.get('Капитальный').get('цементобетон') if result.get(
                #     'Капитальный').get('цементобетон') != 0 else '-'
                # ws[f'{column}12'] = result.get('Капитальный').get('цементобетон') if result.get(
                #     'Капитальный').get('цементобетон') != 0 else '-'
                # ws[f'{column}13'] = result.get('Капитальный').get('асфальтобетон') if result.get(
                #     'Капитальный').get('асфальтобетон') != 0 else '-'
                # ws[f'{column}14'] = result.get('Капитальный').get('щебень/гравий, обр.вяжущий') if result.get(
                #     'Капитальный').get('щебень/гравий, обр.вяжущий') != 0 else '-'

                # ОБЛЕГЧЕННЫЕ

                # ws[f'{column}19'] = result.get('Облегченный').get('асфальтобетон') \
                #     if result.get('Облегченный').get('асфальтобетон') != 0 else '-'
                # ws[f'{column}20'] = result.get('Облегченный').get('органоминеральные') \
                #     if result.get('Облегченный').get('органоминеральные') != 0 else '-'
                # ws[f'{column}21'] = result.get('Облегченный').get('щебеночные (гравийные), обработанные вяжущим') \
                #     if result.get('Облегченный').get('щебеночные (гравийные), обработанные вяжущим') != 0 else '-'

                # ПЕРЕХОДНЫЕ

                # ws[f'{column}26'] = result.get('Переходный').get('Щебеночно - гравийно - песчаные') + result.get(
                #     'Переходный').get('щебень/гравий') \
                #     if result.get('Переходный').get('Щебеночно - гравийно - песчаные') != 0 or result.get(
                #     'Переходный').get('щебень/гравий') != 0 else '-'
                # ws[f'{column}27'] = result.get('Переходный').get(
                #     'Грунт и малопрочные каменные материалы, укрепленные вяжущим') \
                #     if result.get('Переходный').get(
                #     'Грунт и малопрочные каменные материалы, укрепленные вяжущим') != 0 else '-'
                # ws[f'{column}28'] = result.get('Переходный').get(
                #     'Грунт, укрепленный различными вяжущими и местными материалами') \
                #     if result.get('Переходный').get(
                #     'Грунт, укрепленный различными вяжущими и местными материалами') != 0 else '-'
                # ws[f'{column}29'] = result.get('Переходный').get('Булыжный и колотый камень(мостовые)') \
                #     if result.get('Переходный').get('Булыжный и колотый камень(мостовые)') != 0 else '-'

                # НИЗШИЕ

                # ws[f'{column}34'] = result.get('Низший').get('Грунт профилированный') \
                #     if result.get('Низший').get('Грунт профилированный') != 0 else '-'
                # ws[f'{column}35'] = result.get('Низший').get('грунт') \
                #     if result.get('Низший').get('грунт') != 0 else '-'

            counter += 1

        if len(self.data) > 2:
            row = 8
            column = column_tuple[counter]
            ws[f'{column}4'] = 'Итог'
            for key, val in sum_res.items():
                # заполнение итог всех типов и видов покрытий
                for material in sum_res.get(key).values():
                    ws[f'{column}{row}'] = material if material != 0 else '-'
                    row += 1
                row += 4

    def write_11 (self):
        '''
        TODO: таблица 4.4 посмотреть можно ли оптимизировать
        '''
        ws = self.wb['11']
        # заполнение 11 листа
        counter = 0
        column_tuple_4_6 = ('AU', 'AX', 'BA', 'BD', 'BG', 'BJ', 'BM')
        # column_tuple_4_4 = ('B', 'E', 'L', 'S', 'Z')
        n = 16
        res_sum_curves_and_slopes = [0, 0, 0, 0]

        sum_total = {
            'Автопавильоны': 0,
            'Площадки отдыха': 0,
            'парковка': 0,
            'Освещение дорог': 0,
            'Линии технологической связи': 0,
            'кабельные': 0,
            'воздушные': 0,
            'Автобусные остановки': 0,
            'ПСП': 0,
            'Ограждения': 0,
            'Сигнальные столбики': 0
        }
        for k1, v1 in self.data.items():
            if isinstance(v1, str):
                continue
            # 4.4
            slopes_list = v1.get('Кривая', {}).get('Продольный уклон', [])
            dict_counter_and_length_slopes = {
                'IА': [0, 0],
                'IБ': [0, 0],
                'IВ': [0, 0],
                'II': [0, 0],
                'III': [0, 0],
                'IV': [0, 0],
                'V': [0, 0]
            }
            try:
                curves_list = [i for i in v1.get('Кривая', {}).get('R') if i[3] == 146]
            except KeyError as e:
                self.errors.write(f'{self.tip_doc} 11 лист таблица 4.4 oбъект "Кривая" не имеет R')
                # self.msg.information(self.parent, f'{self.tip_doc} 11 лист таблица 4.4',
                #                      'Объект "Кривая" не имеет R',
                #                      QMessageBox.StandardButton.Ok,
                #                      QMessageBox.StandardButton.NoButton)
            try:
                categorys_road_list = v1.get('Граница участка дороги', {}).get('категория а/д', )
            except KeyError as e:
                self.errors.write(
                    f'{self.tip_doc} 11 лист таблица 4.4 Объект "Граница участка дороги" не имеет категории а/д')
                # self.msg.information(self.parent, f'{self.tip_doc} 11 лист таблица 4.4',
                #                      'Объект "Граница участка дороги" не имеет категории а/д',
                #                      QMessageBox.StandardButton.Ok,
                #                      QMessageBox.StandardButton.NoButton)
            dict_counter_and_length_curves = {
                'IА': [0, 0],
                'IБ': [0, 0],
                'IВ': [0, 0],
                'II': [0, 0],
                'III': [0, 0],
                'IV': [0, 0],
                'V': [0, 0]}
            for idx, category in enumerate(categorys_road_list):
                # находим следующий тип покрытия
                if category == categorys_road_list[-1]:
                    break
                elif category == categorys_road_list[0]:
                    next_category = categorys_road_list[1]
                else:
                    next_category = categorys_road_list[idx % len(categorys_road_list) + 1]
                for curve in curves_list:
                    # посчитать количество и протяженность кривых
                    if category[-2] <= curve[-2] <= next_category[-2] and category[-1] <= curve[-1] <= next_category[
                        -1]:
                        if category[0] == 'IА' and 0.0 < abs(float(curve[0])) < 1200.0:
                            dict_counter_and_length_curves['IА'][0] += 1
                            dict_counter_and_length_curves['IА'][1] += curve[1]
                        elif category[0] == 'IБ' and 0.0 < abs(float(curve[0])) < 800.0:
                            dict_counter_and_length_curves['IБ'][0] += 1
                            dict_counter_and_length_curves['IБ'][1] += curve[1]
                        elif category[0] == 'IВ' and 0.0 < abs(float(curve[0])) < 600.0:
                            dict_counter_and_length_curves['IВ'][0] += 1
                            dict_counter_and_length_curves['IВ'][1] += curve[1]
                        elif category[0] == 'II' and 0.0 < abs(float(curve[0])) < 800.0:
                            dict_counter_and_length_curves['II'][0] += 1
                            dict_counter_and_length_curves['II'][1] += curve[1]
                        elif category[0] == 'III' and 0.0 < abs(float(curve[0])) < 600.0:
                            dict_counter_and_length_curves['III'][0] += 1
                            dict_counter_and_length_curves['III'][1] += curve[1]
                        elif category[0] == 'IV' and 0.0 < abs(float(curve[0])) < 300.0:
                            dict_counter_and_length_curves['IV'][0] += 1
                            dict_counter_and_length_curves['IV'][1] += curve[1]
                        elif category[0] == 'V' and 0.0 < abs(float(curve[0])) < 150.0:
                            dict_counter_and_length_curves['V'][0] += 1
                            dict_counter_and_length_curves['V'][1] += curve[1]
                for slope in slopes_list:
                    # посчитать количество и протяженность продольных углов
                    if (category[-2] <= slope[-2] <= next_category[-2] and category[-1] <= slope[-1] <= next_category[
                        -1]):
                        if category[0] == 'IА' and 0.0 < abs(float(slope[0])) > 30:
                            dict_counter_and_length_slopes['IА'][0] += 1
                            dict_counter_and_length_slopes['IА'][1] += slope[1]
                        elif category[0] == 'IБ' and 0.0 < abs(float(slope[0])) > 40:
                            dict_counter_and_length_slopes['IБ'][0] += 1
                            dict_counter_and_length_slopes['IБ'][1] += slope[1]
                        elif category[0] == 'IВ' and 0.0 < abs(float(slope[0])) > 50:
                            dict_counter_and_length_slopes['IВ'][0] += 1
                            dict_counter_and_length_slopes['IВ'][1] += slope[1]
                        elif category[0] == 'II' and 0.0 < abs(float(slope[0])) > 40:
                            dict_counter_and_length_slopes['II'][0] += 1
                            dict_counter_and_length_slopes['II'][1] += slope[1]
                        elif category[0] == 'III' and 0.0 < abs(float(slope[0])) > 50:
                            dict_counter_and_length_slopes['III'][0] += 1
                            dict_counter_and_length_slopes['III'][1] += slope[1]
                        elif category[0] == 'IV' and 0.0 < abs(float(slope[0])) > 60:
                            dict_counter_and_length_slopes['IV'][0] += 1
                            dict_counter_and_length_slopes['IV'][1] += slope[1]
                        elif category[0] == 'V' and 0.0 < abs(float(slope[0])) > 70:
                            dict_counter_and_length_slopes['V'][0] += 1
                            dict_counter_and_length_slopes['V'][1] += slope[1]

            res_sum_curves_and_slopes[0] += sum(i[0] for i in dict_counter_and_length_curves.values())
            res_sum_curves_and_slopes[1] += sum(i[1] for i in dict_counter_and_length_curves.values()) / 1000
            res_sum_curves_and_slopes[2] += sum(i[0] for i in dict_counter_and_length_slopes.values())
            res_sum_curves_and_slopes[3] += sum(i[1] for i in dict_counter_and_length_slopes.values()) / 1000
            if len(self.data) > 2:
                ws.unmerge_cells(start_row = n, start_column = 2, end_row = n, end_column = 4)
                ws.unmerge_cells(start_row = n, start_column = 5, end_row = n, end_column = 11)
                ws.unmerge_cells(start_row = n, start_column = 12, end_row = n, end_column = 18)
                ws.unmerge_cells(start_row = n, start_column = 19, end_row = n, end_column = 25)
                ws.unmerge_cells(start_row = n, start_column = 26, end_row = n, end_column = 33)
                for col in range(2, 34):
                    ws.cell(row = n, column = col).border = self.table_cells_border
                ws.merge_cells(f'B{n}:AG{n}')
                ws[f'B{n}'] = k1
                n += 1

                ws[f'B{n}'] = self.data_interface.get('year', None)
                ws[f'E{n}'] = res_sum_curves_and_slopes[0] if res_sum_curves_and_slopes[0] != 0 else '-'
                ws[f'L{n}'] = res_sum_curves_and_slopes[1] if res_sum_curves_and_slopes[1] != 0 else '-'
                ws[f'S{n}'] = res_sum_curves_and_slopes[2] if res_sum_curves_and_slopes[2] != 0 else '-'
                ws[f'Z{n}'] = res_sum_curves_and_slopes[3] if res_sum_curves_and_slopes[3] != 0 else '-'
                n += 1
            else:
                ws[f'B{n}'] = self.data_interface.get('year', None)
                ws[f'E{n}'] = res_sum_curves_and_slopes[0] if res_sum_curves_and_slopes[0] != 0 else '-'
                ws[f'L{n}'] = res_sum_curves_and_slopes[1] if res_sum_curves_and_slopes[1] != 0 else '-'
                ws[f'S{n}'] = res_sum_curves_and_slopes[2] if res_sum_curves_and_slopes[2] != 0 else '-'
                ws[f'Z{n}'] = res_sum_curves_and_slopes[3] if res_sum_curves_and_slopes[3] != 0 else '-'

            # 4.6
            column = column_tuple_4_6[counter]
            # шапка участки
            if len(self.data) > 2:
                ws[f'{column}6'] = f'Участок {counter + 1} \n {self.data_interface.get("year", None)} г.'
            else:
                ws[f'{column}6'] = f'{self.data_interface.get("year", None)}'

            # автопавильоны капитального типа шт
            sum_stop_bus = sum(1 for i in v1.get('Остановка').get('Наличие павильона') if i[0] == 'да') if v1.get(
                'Остановка', {}).get('Наличие павильона', False) else 0
            sum_total['Автопавильоны'] += sum_stop_bus
            ws[f"{column}14"] = sum_stop_bus if sum_stop_bus != 0 else '-'

            # площадки отдыха шт
            sum_recreation_area = sum(
                1 for i in v1.get('Проезжая часть').get('Назначение') if i[0] == 'площадка отдыха') if v1.get(
                'Проезжая часть', {}).get('Назначение', False) else 0
            sum_total['Площадки отдыха'] += sum_recreation_area
            ws[f"{column}16"] = sum_recreation_area if sum_recreation_area != 0 else '-'

            # площадка для стоянок и остановок автомобилей шт
            sum_parking = sum(1 for i in v1.get('Проезжая часть').get('Назначение') if i[0] == 'парковка') if v1.get(
                'Проезжая часть', {}).get('Назначение', False) else 0
            sum_total['парковка'] += sum_parking
            ws[f"{column}17"] = sum_parking if sum_parking != 0 else '-'

            # освещение дороги км
            sum_light = round(
                sum(float(x[1]) for x in v1.get('Опоры освещения и контактной сети').get('Статус')) / 1000,
                3) if v1.get('Опоры освещения и контактной сети', {}).get('Статус', False) else 0
            sum_total['Освещение дорог'] += sum_light
            ws[f"{column}19"] = sum_light if sum_light != 0 else '-'

            # линии технологической связи кабельные км
            sum_line_communications_cabel = round(sum(float(x[1]) for x in
                                                      v1.get('Подземная комуникация').get('Вид коммуникации')) / 1000,
                                                  3) if (
                v1.get('Подземная комуникация', {}).get('Вид коммуникации', False)) else 0
            sum_total['кабельные'] += sum_line_communications_cabel
            ws[f"{column}23"] = sum_line_communications_cabel if sum_line_communications_cabel != 0 else '-'

            # линии технологической связи воздушные км
            sum_line_communications_air = round(sum(float(x[1]) for x in
                                                    v1.get('Воздушная коммуникация').get('Вид коммуникации')) / 1000,
                                                3) if v1.get(
                'Воздушная коммуникация', {}).get('Вид коммуникации', False) else 0
            sum_total['воздушные'] += sum_line_communications_air
            ws[f"{column}24"] = sum_line_communications_air if sum_line_communications_air != 0 else '-'

            # всего км
            total_line_communications = sum_line_communications_air + sum_line_communications_cabel
            sum_total['Линии технологической связи'] += total_line_communications
            ws[f"{column}20"] = total_line_communications if total_line_communications != 0 else '-'

            # остановки шт
            count_stop_bus = len(v1.get('Остановка', {}).get('Название остановки', []))
            sum_total['Автобусные остановки'] += count_stop_bus
            ws[f"{column}25"] = count_stop_bus if count_stop_bus != 0 else '-'

            # ПСП шт
            sum_psp = sum(1 for i in v1.get('Проезжая часть', {}).get('Назначение', []) if
                          i[0] in ['полоса торможения', 'полоса разгона'])
            sum_total['ПСП'] += sum_psp
            ws[f"{column}26"] = sum_psp if sum_psp != 0 else '-'

            # ограждения км
            sum_fencing = round(sum(float(x[1]) for k in
                                    ['Нестандартное ограждение', 'Пешеходное ограждение', 'Тросовое ограждение',
                                     'Типа Нью-Джерси', 'Металическое барьерное ограждение', 'Парапетное ограждение'
                                     ] for x in
                                    v1.get(k, {}).get('Статус', [])) / 1000, 3)
            sum_total['Ограждения'] += sum_fencing
            ws[f"{column}28"] = sum_fencing if sum_fencing != 0 else '-'

            # сигнальные столбики шт
            count_sign_column = sum(x[4] for x in v1.get('Сигнальные столбики', {}).get('Статус', [])) if v1.get(
                'Сигнальные столбики', {}).get('Статус', []) else 0
            sum_total['Сигнальные столбики'] += count_sign_column
            ws[f"{column}29"] = count_sign_column if count_sign_column != 0 else '-'

            sum_sign = {'всего': 0,
                        'предупреждающие': 0,
                        'приоритета': 0,
                        'запрещающие': 0,
                        'предписывающие': 0,
                        'особых предписаний': 0,
                        'сервиса': 0,
                        'информационные': 0,
                        'дополнительной информации': 0}

            # подсчет знаков шт
            # try:
            for k, v in v1.items():
                if k[0].isdigit():

                    sum_sign['всего'] += len(v['Статус']) if v.get('Статус', []) else 0

                    if k[0] == '1':
                        sum_sign['предупреждающие'] += len(v['Статус']) if v.get('Статус', []) else 0
                    elif k[0] == '2':
                        sum_sign['приоритета'] += len(v['Статус']) if v.get('Статус', []) else 0
                    elif k[0] == '3':
                        sum_sign['запрещающие'] += len(v['Статус']) if v.get('Статус', []) else 0
                    elif k[0] == '4':
                        sum_sign['предписывающие'] += len(v['Статус']) if v.get('Статус', []) else 0
                    elif k[0] == '5':
                        sum_sign['особых предписаний'] += len(v['Статус']) if v.get('Статус', []) else 0
                    elif k[0] == '6':
                        sum_sign['информационные'] += len(v['Статус']) if v.get('Статус', []) else 0
                    elif k[0] == '7':
                        sum_sign['сервиса'] += len(v['Статус']) if v.get('Статус', []) else 0
                    elif k[0] == '8':
                        sum_sign['дополнительной информации'] += len(v['Статус']) if v.get('Статус', []) else 0
            for k, v in sum_sign.items():
                if k in sum_total:
                    sum_total[k] += v
                else:
                    sum_total[k] = v
            # знаки шт
            ws[f'{column}30'] = sum_sign.get('всего') if sum_sign.get('всего') != 0 else '-'
            ws[f'{column}32'] = sum_sign.get('предупреждающие') if sum_sign.get('предупреждающие') != 0 else '-'
            ws[f'{column}33'] = sum_sign.get('приоритета') if sum_sign.get('приоритета') != 0 else '-'
            ws[f'{column}34'] = sum_sign.get('запрещающие') if sum_sign.get('запрещающие') != 0 else '-'
            ws[f'{column}35'] = sum_sign.get('предписывающие') if sum_sign.get('предписывающие') != 0 else '-'
            ws[f'{column}36'] = sum_sign.get('особых предписаний') if sum_sign.get('особых предписаний') != 0 else '-'
            ws[f'{column}37'] = sum_sign.get('информационные') if sum_sign.get('информационные') != 0 else '-'
            ws[f'{column}38'] = sum_sign.get('сервиса') if sum_sign.get('сервиса') != 0 else '-'
            ws[f'{column}39'] = sum_sign.get('дополнительной информации') if sum_sign.get(
                'дополнительной информации') != 0 else '-'
            counter += 1

        if len(self.data) > 2:
            # если участков несколько столбц итого
            column = column_tuple_4_6[counter]
            ws[f'{column}6'] = 'Итог'
            ws[f'{column}14'] = sum_total.get('Автопавильоны') if sum_total.get('Автопавильоны', 0) != 0 else '-'
            ws[f'{column}16'] = sum_total.get('Площадки отдыха') if sum_total.get('Площадки отдыха', 0) != 0 else '-'
            ws[f'{column}17'] = sum_total.get('парковка') if sum_total.get('парковка', 0) != 0 else '-'
            ws[f'{column}19'] = sum_total.get('Освещение дорог') if sum_total.get('Освещение дорог', 0) != 0 else '-'
            ws[f'{column}20'] = sum_total.get('Линии технологической связи') if sum_total.get(
                'Линии технологической связи', 0) != 0 else '-'
            ws[f'{column}23'] = sum_total.get('кабельные') if sum_total.get('кабельные', 0) != 0 else '-'
            ws[f'{column}24'] = sum_total.get('воздушные') if sum_total.get('воздушные', 0) != 0 else '-'
            ws[f'{column}25'] = sum_total.get('Автобусные остановки') if sum_total.get('Автобусные остановки',
                                                                                       0) != 0 else '-'
            ws[f'{column}26'] = sum_total.get('ПСП') if sum_total.get('ПСП', 0) != 0 else '-'
            ws[f'{column}28'] = sum_total.get('Ограждения') if sum_total.get('Ограждения', 0) != 0 else '-'
            ws[f'{column}29'] = sum_total.get('Сигнальные столбики') if sum_total.get('Сигнальные столбики',
                                                                                      0) != 0 else '-'
            ws[f'{column}30'] = sum_total.get('всего') if sum_total.get('всего', 0) != 0 else '-'
            ws[f'{column}32'] = sum_total.get('предупреждающие') if sum_total.get('предупреждающие', 0) != 0 else '-'
            ws[f'{column}33'] = sum_total.get('приоритета') if sum_total.get('приоритета', 0) != 0 else '-'
            ws[f'{column}34'] = sum_total.get('запрещающие') if sum_total.get('запрещающие', 0) != 0 else '-'
            ws[f'{column}35'] = sum_total.get('предписывающие') if sum_total.get('предписывающие', 0) != 0 else '-'
            ws[f'{column}36'] = sum_total.get('особых предписаний') if sum_total.get('особых предписаний',
                                                                                     0) != 0 else '-'
            ws[f'{column}37'] = sum_total.get('информационные') if sum_total.get('информационные', 0) != 0 else '-'
            ws[f'{column}38'] = sum_total.get('сервиса') if sum_total.get('сервиса', 0) != 0 else '-'
            ws[f'{column}39'] = sum_total.get('дополнительной информации') if sum_total.get('дополнительной информации',
                                                                                            0) != 0 else '-'

    def write_12 (self):
        ws = self.wb['12']
        # 4.7.1
        rows_avtovokzal = 11
        rows_gbdd = 11
        rows_sto = 37
        rows_hotels = 37
        for name_district, obj in self.data.items():
            if isinstance(obj, str):
                continue

            for idx, value in enumerate(obj.get('Здание', {}).get('Назначение', [])):

                if value[0] == 'Автостанция/автовокзал':
                    # 4.7.1
                    ws[f'B{rows_avtovokzal}'] = obj.get('Здание', {}).get('Наименование')[idx][0]
                    ws[f'I{rows_avtovokzal}'] = obj.get('Здание', {}).get('Адрес')[idx][0]
                    ws[f'N{rows_avtovokzal}'] = obj.get('Здание', {}).get('Привязка по оси')[idx][0]
                    rows_avtovokzal += 1
                elif value[0] == 'Перецепной КДП/пост ГИБДД':
                    # 4.7.2
                    ws[f'AJ{rows_gbdd}'] = obj.get('Здание', {}).get('Наименование')[idx][0]
                    ws[f'AS{rows_gbdd}'] = obj.get('Здание', {}).get('Адрес')[idx][0]
                    ws[f'BA{rows_gbdd}'] = obj.get('Здание', {}).get('Привязка по оси')[idx][0]
                    rows_gbdd += 1
                elif value[0] == 'Гостиница/мотель/кемпинг':
                    # 4.7.4
                    ws[f'B{rows_sto}'] = obj.get('Здание', {}).get('Наименование')[idx][0]
                    ws[f'I{rows_sto}'] = obj.get('Здание', {}).get('Адрес')[idx][0]
                    ws[f'N{rows_sto}'] = obj.get('Здание', {}).get('Привязка по оси')[idx][0]
                    rows_sto += 1
                elif value[0] == 'СТО':
                    # 4.7.3
                    ws[f'AJ{rows_hotels}'] = obj.get('Здание', {}).get('Адрес')[idx][0]
                    ws[f'AS{rows_hotels}'] = obj.get('Здание', {}).get('Привязка по оси')[idx][0]
                    ws[f'BD{rows_hotels}'] = obj.get('Здание', {}).get('Наименование')[idx][0]
                    rows_hotels += 1

    def write_13 (self):
        ws = self.wb['13']
        rows_azs = 10
        rows_car_wash = 10
        rows_ws = 37
        rows_food = 37
        for name_district, obj in self.data.items():
            if isinstance(obj, str):
                continue

            for idx, value in enumerate(obj.get('Здание', {}).get('Назначение', [])):
                try:
                    if value[0] == 'АЗС':
                        # 4.7.5
                        ws[f'B{rows_azs}'] = obj.get('Здание', {}).get('Адрес')[idx][0] if obj.get('Здание', {}).get(
                            'Адрес', False) else '-'
                        ws[f'K{rows_azs}'] = obj.get('Здание', {}).get('Привязка по оси')[idx][0] if obj.get('Здание',
                                                                                                             {}).get(
                            'Привязка по оси', False) else '-'
                        ws[f'V{rows_azs}'] = obj.get('Здание', {}).get('Наименование')[idx][0] if obj.get('Здание',
                                                                                                          {}).get(
                            'Наименование', False) else '-'
                        rows_azs += 1
                    elif value[0] == 'Моечный пункт':
                        # 4.7.6
                        ws[f'AJ{rows_car_wash}'] = obj.get('Здание', {}).get('Адрес')[idx][0] if obj.get('Здание',
                                                                                                         {}).get(
                            'Адрес', False) else '-'
                        ws[f'AS{rows_car_wash}'] = obj.get('Здание', {}).get('Привязка по оси')[idx][0] if obj.get(
                            'Здание',
                            {}).get(
                            'Привязка по оси', False) else '-'
                        ws[f'BD{rows_car_wash}'] = obj.get('Здание', {}).get('Наименование')[idx][0] if obj.get(
                            'Здание',
                            {}).get(
                            'Наименование', False) else '-'
                        rows_car_wash += 1
                    elif value[0] == 'Общественный туалет':
                        # 4.7.7
                        ws[f'B{rows_ws}'] = obj.get('Здание', {}).get('Адрес')[idx][0] if obj.get('Здание', {}).get(
                            'Адрес', False) else '-'
                        ws[f'I{rows_ws}'] = obj.get('Здание', {}).get('Привязка по оси')[idx][0] if obj.get('Здание',
                                                                                                            {}).get(
                            'Привязка по оси', False) else '-'
                        ws[f'N{rows_ws}'] = obj.get('Здание', {}).get('Наименование')[idx][0] if obj.get('Здание',
                                                                                                         {}).get(
                            'Наименование', False) else '-'
                        rows_ws += 1
                    elif value[0] == 'Ресторан/кафе/столовая':
                        # 4.7.8
                        ws[f'AJ{rows_food}'] = obj.get('Здание', {}).get('Наименование')[idx][0] if obj.get('Здание',
                                                                                                            {}).get(
                            'Наименование', False) else '-'
                        ws[f'AQ{rows_food}'] = obj.get('Здание', {}).get('Адрес')[idx][0] if obj.get('Здание', {}).get(
                            'Адрес', False) else '-'
                        ws[f'AV{rows_food}'] = obj.get('Здание', {}).get('Привязка по оси')[idx][0] if \
                            obj.get('Здание', {}).get('Привязка по оси', False) else '-'
                except KeyError as e:
                    self.errors.write(f'Объект {value[0]}({value[-2]}) некорректно заполнен \n')

                except AttributeError as e:
                    self.errors.write(f'не могу заполнить {value[0]} строка {rows_food} \n')

                rows_food += 1

    def write_14 (self):
        ws = self.wb['14']
        rows_medical = 8
        for name_district, obj in self.data.items():
            if isinstance(obj, str):
                continue
            for idx, value in enumerate(obj.get('Здание', {}).get('Назначение', [])):
                if value[0] == 'Пункт первой медпомощи/почта/телефон':
                    # 4.7.5
                    ws[f'B{rows_medical}'] = obj.get('Здание', {}).get('Адрес')[idx][0] if obj.get('Здание', {}).get(
                        'Адрес', False) else '-'
                    ws[f'O{rows_medical}'] = obj.get('Здание', {}).get('Привязка по оси')[idx][0] if obj.get('Здание',
                                                                                                             {}).get(
                        'Привязка по оси', False) else '-'
                    ws[f'Y{rows_medical}'] = obj.get('Здание', {}).get('Наименование')[idx][0] if obj.get('Здание',
                                                                                                          {}).get(
                        'Наименование', False) else '-'
                    rows_medical += 1

    def write_17 (self):
        """
        :return:
        """
        ws = self.wb['17']
        counter = 0

        columns_left = ('J', 'O', 'T', 'Y', 'AD')
        cells_left = ('L', 'Q', 'V', 'AA', 'AF')
        columns_right = ('AS', 'AX', 'BC', 'BH', 'BM')
        cells_right = ('AU', 'AZ', 'BE', 'BJ', 'BO')
        total_sum_4_10_2 = {}
        total_sum_4_10_3 = {}
        total_sum_4_10_4 = {}
        total_sum_4_10_5 = {}

        def count_4_10_2 (column, cell, list_obj):
            # 4.10.2 Сводная ведомость наличия тоннелей, галерей и пешеходных переходов в разных уровнях
            types_of_structures = {
                "Тоннель (галерея)": [0, 0],
                "Пешеходный переход подземный": [0, 0],
                "Пешеходный переход надземный": [0, 0],
            }
            sum_crosswalk = [0, 0]
            sum_all = [0, 0]

            for key, value in types_of_structures.items():

                # 4.10.2
                result = list_obj.get(key)
                if result is not None:
                    types_of_structures.get(key)[0] += 1
                    types_of_structures.get(key)[1] += result.get(list(result.keys())[0])[0][2]
                if key == 'Тоннель (галерея)':
                    ws[f'{column}14'] = types_of_structures.get(key)[0] if types_of_structures.get(key)[
                                                                               0] != 0 else '-'
                    ws[f'{cell}14'] = types_of_structures.get(key)[1] if types_of_structures.get(key)[
                                                                             1] != 0 else '-'

                elif key == 'Пешеходный переход надземный':
                    ws[f'{column}19'] = types_of_structures.get(key)[0] if types_of_structures.get(key)[
                                                                               0] != 0 else '-'
                    ws[f'{cell}19'] = types_of_structures.get(key)[1] if types_of_structures.get(key)[
                                                                             1] != 0 else '-'
                    sum_crosswalk[0] += types_of_structures.get(key)[0] if types_of_structures.get(key)[
                                                                               0] != 0 else 0
                    sum_crosswalk[1] += types_of_structures.get(key)[1] if types_of_structures.get(key)[
                                                                               1] != 0 else 0
                elif key == 'Пешеходный переход подземный':
                    ws[f'{column}20'] = types_of_structures.get(key)[0] if types_of_structures.get(key)[
                                                                               0] != 0 else '-'
                    ws[f'{cell}20'] = types_of_structures.get(key)[1] if types_of_structures.get(key)[
                                                                             1] != 0 else '-'
                    sum_crosswalk[0] += types_of_structures.get(key)[0] if types_of_structures.get(key)[
                                                                               0] != 0 else 0
                    sum_crosswalk[1] += types_of_structures.get(key)[1] if types_of_structures.get(key)[
                                                                               1] != 0 else 0
                sum_all[0] += types_of_structures.get(key)[0]
                sum_all[1] += types_of_structures.get(key)[1]
                # количество пешеходных переходов(суммарно)
                ws[f'{column}16'] = sum_crosswalk[0] if sum_crosswalk[0] != 0 else '-'
                ws[f'{cell}16'] = sum_crosswalk[1] if sum_crosswalk[1] != 0 else '-'
            # всего пешеходных ограждений
            ws[f'{column}21'] = sum_all[0] if sum_all[0] != 0 else '-'
            ws[f'{cell}21'] = sum_all[1] if sum_all[1] != 0 else '-'
            return types_of_structures

        def count_4_10_3 (column, cell, list_obj):
            # 4.10.3 трубы
            row = 37
            sum_all = [0, 0]
            pipes = {
                "Металлические": [0, 0],
                "Железобетонные": [0, 0],
                "Бетоннометаллические": [0, 0],  # нет в свпд
                "Каменные": [0, 0],
                "Деревянные": [0, 0],
                "Асбестоцементные": [0, 0],
            }
            for lst in list_obj.get('Водопропускная труба', {}).get('Материал', []):
                if lst[0] == 'металл':
                    pipes.get('Металлические')[0] += 1
                    pipes.get('Металлические')[1] += lst[1]
                elif lst[0] == 'ж/б':
                    pipes.get('Железобетонные')[0] += 1
                    pipes.get('Железобетонные')[1] += lst[1]
                elif lst[0] == 'камень':
                    pipes.get('Каменные')[0] += 1
                    pipes.get('Каменные')[1] += lst[1]
                elif lst[0] == 'дерево':
                    pipes.get('Деревянные')[0] += 1
                    pipes.get('Деревянные')[1] += lst[1]
                elif lst[0] == 'асбоцемент':
                    pipes.get('Асбестоцементные')[0] += 1
                    pipes.get('Асбестоцементные')[1] += lst[1]

            sum_all[0] = len(list_obj.get('Водопропускная труба', {}).get('Материал', []))
            sum_all[1] = sum(i[1] for i in list_obj.get('Водопропускная труба', {}).get('Материал', []))
            for value in pipes.values():
                ws[f'{column}{row}'] = value[0] if value[0] != 0 else '-'
                ws[f'{cell}{row}'] = value[1] if value[1] != 0 else '-'
                row += 1

            ws[f'{column}{row}'] = sum_all[0] if sum_all[0] != 0 else '-'
            ws[f'{cell}{row}'] = sum_all[1] if sum_all[1] != 0 else '-'
            return pipes

        def count_4_10_4 (column, cell, list_obj):
            # 4.10.4 паромные переправы
            row = 14
            ferry_crossings = {'самоходные': [0, 0],
                               'буксирные': [0, 0],
                               'канатные': [0, 0],
                               }
            sum_all = [0, 0]

            for lst in list_obj.get('Переправа паромная (ледовая)', {}).get('Способ передвижения парома', []):
                if lst[0] == 'Самоходные':
                    ferry_crossings.get('самоходные')[0] += 1
                    ferry_crossings.get('самоходные')[1] += lst[1]
                elif lst[0] == 'Буксирные':
                    ferry_crossings.get('буксирные')[0] += 1
                    ferry_crossings.get('буксирные')[1] += lst[1]
                elif lst[0] == 'Канатные':
                    ferry_crossings.get('канатные')[0] += 1
                    ferry_crossings.get('канатные')[1] += lst[1]
            sum_all[0] = len(list_obj.get('Переправа паромная (ледовая)', {}).get('Способ передвижения парома', []))
            sum_all[1] = sum(
                i[1] for i in list_obj.get('Переправа паромная (ледовая)', {}).get('Способ передвижения парома', []))
            for val in ferry_crossings.values():
                ws[f'{column}{row}'] = val[0] if val[0] != 0 else '-'
                ws[f'{cell}{row}'] = val[1] if val[1] != 0 else '-'
                row += 1
            ws[f'{column}{row}'] = sum_all[0] if sum_all[0] != 0 else '-'
            ws[f'{cell}{row}'] = sum_all[1] if sum_all[1] != 0 else '-'
            return ferry_crossings

        def count_4_10_5 (column, cell, list_obj):
            # 4.10.5 подпорные стенки
            row = 37
            retaining_walls = {
                'ж\б': [0, 0],
                'дерево': [0, 0],
                'камень': [0, 0],
                'бетон': [0, 0]
            }
            sum_all = [0, 0]
            for lst in list_obj.get('Подпорная стенка', {}).get('Материал', []):
                if lst[0] == 'ж/б':
                    retaining_walls.get('ж\б')[0] += 1
                    retaining_walls.get('ж\б')[1] += lst[1]
                elif lst[0] == 'камень':
                    retaining_walls.get('камень')[0] += 1
                    retaining_walls.get('камень')[1] += lst[1]
                elif lst[0] == 'дерево':
                    retaining_walls.get('дерево')[0] += 1
                    retaining_walls.get('дерево')[1] += lst[1]
                elif lst[0] == 'бетон':
                    retaining_walls.get('бетон')[0] += 1
                    retaining_walls.get('бетон')[1] += lst[1]
            sum_all[0] = len(list_obj.get('Подпорная стенка', {}).get('Материал', []))
            sum_all[1] = sum(i[1] for i in list_obj.get('Подпорная стенка', {}).get('Материал', []))
            for val in retaining_walls.values():
                ws[f'{column}{row}'] = val[0] if val[0] != 0 else '-'
                ws[f'{cell}{row}'] = val[1] if val[1] != 0 else '-'
                row += 1
            ws[f'{column}{row}'] = sum_all[0] if sum_all[0] != 0 else '-'
            ws[f'{cell}{row}'] = sum_all[1] if sum_all[1] != 0 else '-'
            return retaining_walls

        for k, v in self.data.items():
            if isinstance(v, str):
                continue

            column_left = columns_left[counter]
            cell_left = cells_left[counter]
            column_right = columns_right[counter]
            cell_right = cells_right[counter]
            if len(self.data) > 2:
                ws[f'{column_left}6'] = f'{k.title()} \n {self.data_interface.get("year", None)} г.'
                ws[f'{column_left}29'] = f'{k.title()} \n {self.data_interface.get("year", None)} г.'
                ws[f'{column_right}6'] = f'{k.title()} \n {self.data_interface.get("year", None)} г.'
                ws[f'{column_right}29'] = f'{k.title()} \n {self.data_interface.get("year", None)} г.'

            else:
                ws[f'{column_left}6'] = f'{self.data_interface.get("year", None)}'
                ws[f'{column_left}29'] = f'{self.data_interface.get("year", None)}'
                ws[f'{column_right}6'] = f'{self.data_interface.get("year", None)}'
                ws[f'{column_right}29'] = f'{self.data_interface.get("year", None)}'

            types_of_structures = count_4_10_2(column_left, cell_left, v)
            for key, val in types_of_structures.items():
                if key in total_sum_4_10_2:
                    total_sum_4_10_2[key][0] += val[0]
                    total_sum_4_10_2[key][1] += val[1]
                else:
                    total_sum_4_10_2[key] = val
            pipes = count_4_10_3(column_left, cell_left, v)
            for key, val in pipes.items():
                if key in total_sum_4_10_3:
                    total_sum_4_10_3[key][0] += val[0]
                    total_sum_4_10_3[key][1] += val[1]
                else:
                    total_sum_4_10_3[key] = val
            ferry_crossing = count_4_10_4(column_right, cell_right, v)
            for key, val in ferry_crossing.items():
                if key in total_sum_4_10_4:
                    total_sum_4_10_4[key][0] += val[0]
                    total_sum_4_10_4[key][1] += val[1]
                else:
                    total_sum_4_10_4[key] = val
            walls = count_4_10_5(column_right, cell_right, v)
            for key, val in walls.items():
                if key in total_sum_4_10_5:
                    total_sum_4_10_5[key][0] += val[0]
                    total_sum_4_10_5[key][1] += val[1]
                else:
                    total_sum_4_10_5[key] = val
            counter += 1
        # итого
        if len(self.data) > 2:
            column_left = columns_left[counter]
            cell_left = cells_left[counter]
            column_right = columns_right[counter]
            cell_right = cells_right[counter]
            # 4.10.2
            ws[f'{column_left}6'] = 'Итого'
            ws[f'{column_left}14'] = total_sum_4_10_2.get('Тоннель (галерея)', [0, 0])[0] if \
                total_sum_4_10_2.get('Тоннель (галерея)', [0, 0])[0] != 0 else '-'
            ws[f'{cell_left}14'] = total_sum_4_10_2.get('Тоннель (галерея)', [0, 0])[1] if \
                total_sum_4_10_2.get('Тоннель (галерея)', [0, 0])[1] != 0 else '-'
            ws[f'{column_left}16'] = total_sum_4_10_2.get('Пешеходный переход надземный', [0, 0])[0] + \
                                     total_sum_4_10_2.get('Пешеходный переход подземный', [0, 0])[0] if \
                total_sum_4_10_2.get('Пешеходный переход надземный', [0, 0])[0] + \
                total_sum_4_10_2.get('Пешеходный переход подземный', [0, 0])[0] != 0 else '-'
            ws[f'{cell_left}16'] = total_sum_4_10_2.get('Пешеходный переход надземный', [0, 0])[1] + \
                                   total_sum_4_10_2.get('Пешеходный переход подземный', [0, 0])[1] if \
                total_sum_4_10_2.get('Пешеходный переход надземный', [0, 0])[1] + \
                total_sum_4_10_2.get('Пешеходный переход подземный', [0, 0])[1] != 0 else '-'
            ws[f'{column_left}19'] = total_sum_4_10_2.get('Пешеходный переход надземный', [0, 0])[0] if \
                total_sum_4_10_2.get('Пешеходный переход надземный', [0, 0])[0] != 0 else '-'
            ws[f'{cell_left}19'] = total_sum_4_10_2.get('Пешеходный переход надземный', [0, 0])[1] if \
                total_sum_4_10_2.get('Пешеходный переход надземный', [0, 0])[1] != 0 else '-'
            ws[f'{column_left}20'] = total_sum_4_10_2.get('Пешеходный переход подземный', [0, 0])[0] if \
                total_sum_4_10_2.get('Пешеходный переход подземный', [0, 0])[0] != 0 else '-'
            ws[f'{cell_left}20'] = total_sum_4_10_2.get('Пешеходный переход подземный', [0, 0])[1] if \
                total_sum_4_10_2.get('Пешеходный переход подземный', [0, 0])[1] != 0 else '-'
            sum_all = [sum(i[0] for i in total_sum_4_10_2.values()), sum(i[1] for i in total_sum_4_10_2.values())]
            ws[f'{column_left}21'] = sum_all[0]
            ws[f'{cell_left}21'] = sum_all[1]
            # 4.10.3
            ws[f'{column_left}29'] = f'Итого'
            row = 37
            for k, v in total_sum_4_10_3.items():
                ws[f'{column_left}{row}'] = v[0] if v[0] != 0 else '-'
                ws[f'{cell_left}{row}'] = v[1] if v[1] != 0 else '-'
                row += 1
            total_sum_4_10_3 = [sum(i[0] for i in total_sum_4_10_3.values()),
                                sum(i[1] for i in total_sum_4_10_3.values())]
            ws[f'{column_left}{row}'] = total_sum_4_10_3[0] if total_sum_4_10_3[0] != 0 else '-'
            ws[f'{cell_left}{row}'] = total_sum_4_10_3[1] if total_sum_4_10_3[1] != 0 else '-'
            # 4.10.4
            ws[f'{column_right}6'] = 'Итого'
            row = 14
            for k, v in total_sum_4_10_4.items():
                ws[f'{column_right}{row}'] = v[0] if v[0] != 0 else '-'
                ws[f'{cell_right}{row}'] = v[1] if v[1] != 0 else '-'
                row += 1
            total_sum_4_10_4 = [sum(i[0] for i in total_sum_4_10_4.values()),
                                sum(i[1] for i in total_sum_4_10_4.values())]
            ws[f'{column_right}{row}'] = total_sum_4_10_4[0] if total_sum_4_10_4[0] != 0 else '-'
            ws[f'{cell_right}{row}'] = total_sum_4_10_4[1] if total_sum_4_10_4[1] != 0 else '-'

            # 4.10.5
            ws[f'{column_right}29'] = f'Итого'
            row = 37
            for k, v in total_sum_4_10_5.items():
                ws[f'{column_right}{row}'] = v[0] if v[0] != 0 else '-'
                ws[f'{cell_right}{row}'] = v[1] if v[1] != 0 else '-'
                row += 1
            total_sum_4_10_5 = [sum(i[0] for i in total_sum_4_10_5.values()),
                                sum(i[1] for i in total_sum_4_10_5.values())]
            ws[f'{column_right}{row}'] = total_sum_4_10_5[0] if total_sum_4_10_5[0] != 0 else '-'
            ws[f'{cell_right}{row}'] = total_sum_4_10_5[1] if total_sum_4_10_5[1] != 0 else '-'

    def write_18 (self):
        """
            TODO  4.10.6  не реализовано
            Описиваем данные по 18 листу
            :return:
        """
        ws = self.wb['18']
        counter = 0
        counter2 = 1
        count_distr = len(self.data)

        column_tuple_right = ('AP', 'AU', 'AZ', 'BE', 'BJ', 'BM')
        cells_right = ('AP', 'AR', 'AU', 'AW', 'AZ', 'BB', 'BE', 'BG', 'BJ', 'BL')
        column_tuple_left = ('I', 'L', 'O', 'R', 'U', 'X', 'AA', 'AD')

        res_sum_4_10_6 = [0, 0]
        res_sum_4_10_7 = [0, 0]
        res_sum_4_10_9 = {}
        res_sum_4_10_8 = {}

        for k1, v1 in self.data.items():
            total_shoulders = 0
            if isinstance(v1, str):
                continue

            column_l = column_tuple_left[counter]
            column_r = column_tuple_right[counter]
            # заполнение шапки всех таблиц
            if count_distr > 2:
                ws[f'{column_r}6'] = f'{k1.title()} \n {self.data_interface.get("year", None)} г.'
                ws[f'{column_l}6'] = f'{k1.title()} \n {self.data_interface.get("year", None)} г.'
                ws[f'{column_r}30'] = f'{k1.title()} \n {self.data_interface.get("year", None)} г.'
                ws[f'{column_l}30'] = f'{k1.title()} \n {self.data_interface.get("year", None)} г.'

            else:
                ws[f'{column_r}6'] = f'{self.data_interface.get("year", None)}'
                ws[f'{column_l}6'] = f'{self.data_interface.get("year", None)}'
                ws[f'{column_r}30'] = f'{self.data_interface.get("year", None)}'
                ws[f'{column_l}30'] = f'{self.data_interface.get("year", None)}'

            # 4.10.6
            forest_line = v1.get('Лесополоса', {}).get('Тип', [])

            dict_forest_belt = {'': 0,
                                '': 0,
                                '': 0,
                                '': 0,
                                '': 0, }
            decorative_forest_line = []

            # 4.10.7 Сводная ведомость тротуаров и пешеходных дорожек
            sidewalk = v1.get('Тротуар', {}).get('Тип', [])
            sum_sidewalk = 0
            sum_pedestrian_path = 0
            for tip in sidewalk:
                if tip[0] == 'Тротуар':
                    sum_sidewalk += ((tip[-1][0] - tip[-2][0]) * 1000 + (tip[-1][1] - tip[-2][1])) / 1000
                elif tip[0] == 'Пешеходная дорожка':
                    sum_pedestrian_path += ((tip[-1][0] - tip[-2][0]) * 1000 + (tip[-1][1] - tip[-2][1])) / 1000

            ws[f'{column_l}36'] = round(sum_sidewalk, 3) if sum_sidewalk != 0 else '-'
            ws[f'{column_l}37'] = round(sum_pedestrian_path, 3) if sum_pedestrian_path != 0 else '-'
            ws[f'{column_l}39'] = round(sum_sidewalk + sum_pedestrian_path, 3) if (
                                                                                          sum_sidewalk + sum_pedestrian_path) != 0 else '-'
            res_sum_4_10_7[0] += sum_sidewalk
            res_sum_4_10_7[1] += sum_pedestrian_path

            # 4.10.8 Сводная ведомость укрепления обочин
            reinforced_shoulders = v1.get('Укрепленная часть обочины', {}).get('Тип покрытия', [])

            dict_shoulders = {
                'щебень': 0,
                "грунтощебень": 0,
                "асфальтогранулобетон": 0,
                "цементобетон": 0,
                "щебень обработанный вяжущими": 0,
                "гравий": 0,
                "асфальтобетон": 0,
                "грунтогравий": 0,
                "гравий обработанный вяжущими": 0,
                "грунт укрепленный вяжущими": 0,
                "битумоминеральные смеси": 0,
                "другие местные и несвязанные материалы": 0
            }

            for lst in reinforced_shoulders:
                dict_shoulders[lst[0].lower()] += lst[1]
                total_shoulders += lst[1]

            ws[f'{column_r}12'] = total_shoulders / 1000 if total_shoulders != 0 else '-'
            for i in range(15, 26):
                key = list(dict_shoulders.keys())[i - 15]
                if i == 20:
                    i = 21
                ws[f'AI{i}'] = key
                ws[f'{column_r}{i}'] = dict_shoulders[key] / 1000 if dict_shoulders[key] != 0 else '-'

            for k, v in dict_shoulders.items():
                if k in res_sum_4_10_8:
                    res_sum_4_10_8[k] += v
                else:
                    res_sum_4_10_8.update({k: v})

            # 4.10.9 Сводная ведомость съездов (въездов)

            types = {
                "Асфальтобетонные": [0, 0],
                "Цементобетонные": [0, 0],
                "Тротуарная плитка": [0, 0],
                "Щебеночные": [0, 0],
                "Грунтовые": [0, 0],
                "Ж/б плиты": [0, 0],
                "Брусчатка": [0, 0],
                "Булыжник": [0, 0]
            }

            cell = cells_right[counter2]

            for lst in v1.get('Съезд', {}).get('Тип покрытия', []):
                if lst[0].lower() == 'асфальтобетон':
                    types.get('Асфальтобетонные')[0] += 1
                    types.get('Асфальтобетонные')[1] += lst[2]
                elif lst[0].lower() == 'цементобетон':
                    types.get('Цементобетонные')[0] += 1
                    types.get('Цементобетонные')[1] += lst[2]
                elif lst[0] == 'тротуарная плитка':
                    types.get('Тротуарная плитка')[0] += 1
                    types.get('Тротуарная плитка')[1] += lst[2]
                elif 'щебень' in lst[0].lower():
                    types.get('Щебеночные')[0] += 1
                    types.get('Щебеночные')[1] += lst[2]
                elif 'грунт' in lst[0].lower():
                    types.get('Грунтовые')[0] += 1
                    types.get('Грунтовые')[1] += lst[2]
                elif lst[0].lower() == 'ж/б плиты':
                    types.get('Ж/б плиты')[0] += 1
                    types.get('Ж/б плиты')[1] += lst[2]
                elif lst[0].lower() == 'брусчатка':
                    types.get('Брусчатка')[0] += 1
                    types.get('Брусчатка')[1] += lst[2]
                elif lst[0].lower() == 'булыжник':
                    types.get('Булыжник')[0] += 1
                    types.get('Булыжник')[1] += lst[2]

            row = 36
            for k, v in types.items():
                if k in res_sum_4_10_9:
                    res_sum_4_10_9[k][0] += v[0]
                    res_sum_4_10_9[k][1] += v[1]
                else:
                    res_sum_4_10_9.update({k: v})
                ws[f'{column_r}{row}'] = v[0] if v[0] != 0 else '-'
                ws[f'{cell}{row}'] = round(v[1], 2) if v[1] != 0 else '-'
                row += 1
            sum_piece, sum_area = sum(i[0] for i in types.values()), round(sum(i[1] for i in types.values()), 2)
            ws[f'{column_r}44'] = sum_piece
            ws[f'{cell}44'] = sum_area
            counter += 1
            counter2 += 2

        if count_distr > 2:
            column_l = column_tuple_left[counter]
            column_r = column_tuple_right[counter]
            cell = cells_right[counter2]

            # итог 4_10_8
            res = sum(res_sum_4_10_8.values()) / 1000
            ws[f'{column_r}6'] = 'Итого'
            ws[f'{column_r}12'] = res if res != 0 else '-'
            for i in range(15, 26):
                if i == 20:
                    continue
                key = ws[f'AI{i}'].value
                ws[f'{column_r}{i}'] = res_sum_4_10_8.get(key) / 1000 if res_sum_4_10_8.get(key, 0) != 0 else '-'

            # итог 4_10_7
            ws[f'{column_l}30'] = 'Итого'
            ws[f'{column_l}36'] = res_sum_4_10_7[0]
            ws[f'{column_l}37'] = res_sum_4_10_7[1]
            sum_all_4_10_7 = sum(res_sum_4_10_7)
            ws[f'{column_l}39'] = sum_all_4_10_7 if sum_all_4_10_7 != 0 else '-'
            # итог 4_10_9
            row = 36
            ws[f'{column_r}30'] = 'Итого'

            for k, v in res_sum_4_10_9.items():
                ws[f'{column_r}{row}'] = v[0] if v[0] != 0 else '-'
                ws[f'{cell}{row}'] = round(v[1], 2) if v[1] != 0 else '-'
                row += 1

            sum_piece, sum_area = sum(i[0] for i in res_sum_4_10_9.values()), round(
                sum(i[1] for i in res_sum_4_10_9.values()), 2)
            ws[f'{column_r}44'] = sum_piece if sum_piece != 0 else '-'
            ws[f'{cell}44'] = sum_area if sum_area != 0 else '-'


class WriterExcelDAD(WriterExcel):
    #key: категория дороги
    #index:0 Ширина полосы движения, м
    #index:1 Ширина обочины, м
    #index:2 Расчетная скорость, км / ч
    #index:3 Наименьший радиус кривой в плане, м
    #index:4 Наибольший продольный уклон, ‰
    #index:5 Продольная ровность, м / км, не более
    #index:6 Требуемый модуль упругости, Мпа, не менее
    #index:7 Ширина укрепленной обочины
    DICT_NORMATIVE_VALUE = {'IА': (3.75, 3.75, 150, 1200, 30, 4, 230, None),
                            'IБ': (3.75, 3.75, 120, 800, 40, 4, 230, None),
                            'IВ': (3.75, 3.5, 100, 600, 50, 4.5, 230, None),
                            'II': (3.75, 3.5, 120, 800, 40, 4.5, 220, 2),
                            'III': (3.5, 2.5, 100, 600, 50, 5, 200, 1.5),
                            'IV': (3, 2, 80, 300, 60, 6.5, 150, 1),
                            'IVА-р': (3, 2, 80, 265, 60, 6.5, 200, 1),
                            'IVБ-р': (3, 1.5, 60, 125, 70, 6.5, 200, 0.75),
                            'IVА-п': (3, 2, 70, 185, 60, 6.5, 150, 0.75),
                            'IVБ-п': (3, 2, 60, 125, 70, 6.5, 150, 0.75),
                            'V': (4.5, 1.75, 60, 150, 70, 7.5, 100, None),
                            'VА': (4.5, 1.5, 50, 85, 70, 7.5, 100, None),
                            'VБ': (4.5, 1.5, 40, 50, 80, 7.5, None, None),
                            }
    # key: категория дороги
    #index:0 Капитальный
    #index:1 Облегченный
    #index:2 Переходный
    #index:3 Низший
    IRI = {
        'IА': (4, None, None, None),
        'IБ': (4, None, None, None),
        'IВ': (4.5, None, None, None),
        'II': (4.5, None, None, None),
        'III': (5, 5.5, 5.5, None),
        'IV': (6, 6.5, 6.5, None),
        'IVА-р': (6, 6.5, 6.5, None),
        'IVБ-р': (6, 6.5, 6.5, None),
        'IVА-п': (6, 6.5, 6.5, None),
        'IVБ-п': (6, 6.5, 6.5, None),
        'V': (None, 7.5, 8, None),
        'VА': (None, 7.5, 8, None),
        'VБ': (None, 7.5, 8, None)
    }

    def __init__ (self, data: dict = None, path = None, data_interface = None):
        super().__init__(data, path_template_excel = path_template_excel_dad, path = path,
                         data_interface = data_interface)
        self.tip_doc = 'Диагностика'
        self.total_cells_aligment = Alignment(horizontal = 'right', vertical = 'center')
        self.run()
        self.save_file()

    def run(self):
        # self.write_titular()
        # self.write_note()
        # self.write_roadway_width()
        # self.write_shoulder_width()
        self.write_curves()

    def set_chart_text_style (self, chart):
        '''изменяет расмер и стиль текста на диаграмме

        :param chart:  объект диаграмма
        :return:
        '''
        # стиль шрифт
        font = FontText(typeface = 'Times New Roman')
        # создаём загаловок
        cp_title = CharacterProperties(sz = 1400, latin = font)
        cp_legend = CharacterProperties(sz = 1200, latin = font)
        cp_other = CharacterProperties(sz = 1000, latin = font)
        pp_legend = ParagraphProperties(defRPr = cp_legend)
        pp_title = ParagraphProperties(defRPr = cp_title)
        pp_other = ParagraphProperties(defRPr = cp_other)
        rt_y = RichText(p = [Paragraph(pPr = pp_other, endParaRPr = cp_other)])
        rt_x = RichText(p = [Paragraph(pPr = pp_other, endParaRPr = cp_other)],bodyPr = RichTextProperties(vert = 'vert270'))
        for para in chart.title.tx.rich.paragraphs:
            para.pPr = pp_title
        # задаем шрифт текста заголовка оси Y
        if isinstance(chart, BarChart):
            chart.y_axis.title.tx.rich.p[0].pPr = pp_other
            chart.x_axis.txPr = rt_x
            chart.y_axis.txPr = rt_y

            chart.x_axis.txPr.properties.vert ='vert270'
        else:
            chart.dataLabels.txPr = RichText(p = [Paragraph(pPr = pp_title, endParaRPr = cp_title)])

        chart.legend.txPr = RichText(p = [Paragraph(pPr = pp_legend, endParaRPr = cp_legend)])
        chart.title.tx.rich.paragraphs[0].properties.rPr = pp_other

    def write_titular (self):
        ws = self.wb['Титульный']
        ws['A2'] = self.data_interface.get('client', 'Заказчик')
        ws['B15'] = self.data_interface.get('client', 'Заказчик')
        ws['A19'] = f"составлена на {self.data_interface.get('year', '2024')} г."
        ws['A21'] = "Шифр: " + self.data_interface.get('cypher', 'шифр')
        ws['A28'] = self.data_interface.get('contractor', 'Подрядчик')
        ws['A31'] = self.data_interface.get("position_contractor",
                                            "Должность подрядчика") + ' ' + self.data_interface.get("fio_contractor",
                                                                                                    "ФИО подрядчика")
        ws['F28'] = self.data_interface.get('client', 'Заказчик')
        ws['F31'] = self.data_interface.get("position_client", "Должность заказчика") + ' ' + self.data_interface.get(
            "fio_client", "ФИО заказчика")
        ws['A40'] = f"г. Омск {self.data_interface.get('year', '2024')} г."

    def write_scheme (self):
        ws = self.wb['Схема']  # выбираем лист

        try:
            schema = Image(f"{self.path_dir}\Схема.png")
            # schema.width = 1380
            # schema.height = 800
            ws.add_image(schema, 'A2')

        except FileNotFoundError:
            self.errors.write(f'{self.tip_doc} Cхема: Файл Схема.png отсутствует в {self.path_dir}\n')

    def write_note (self):
        '''
        TODO: проверить на работоспособность
        заполенение листа записка
        '''

        ws = self.wb['Записка']
        ws['A3'] = (f'Силами ООО «Сибирь-Регион» были выполнены работы по диагностике (обследованию технического '
                    f'состояния) автомобильной дороги "{self.data["название дороги"]}", с использованием оборудования:')
        ws['C12'] = ''  # какую дату писать
        count_distr = len(self.data)
        i = 14
        row = 24

        for key, value in self.data.items():
            if isinstance(value, str):
                continue
            start, end = value.get('Ось дороги', {}).get('Начало трассы', [])[0][-2:]
            start_str, end_str = self.change_start_and_end_obj(start, end)
            if count_distr > 2:
                ws[f'A{i}'] = f"Начало {key}: {start_str} км"
                ws[f'A{i + 1}'] = f"Конец {key}: {end_str} км"
                i += 2
            else:
                ws['A14'] = f"Начало дороги: {start_str} км"
                ws['A15'] = f"Конец дороги: {end_str} км"
            print(value.get('Граница участка дороги', {}))
            tuple_cateregory = (value.get('Граница участка дороги', {}).get('категория а/д'),)
            tuple_coating = (value.get('Граница участка дороги', {}).get('вид покрытия'),)

            last_cat = tuple_cateregory[0]
            last_coating = tuple_coating[0]
            ws[f'A{row}'] = last_cat[-2][0]
            ws[f'B{row}'] = last_cat[-2][1]
            ws[f'C{row}'] = tuple_cateregory[1][-1][0]
            ws[f'D{row}'] = tuple_cateregory[1][-1][1]
            ws[f'E{row}'] = last_cat[0]
            ws[f'I{row}'] = last_coating[0]
            ws[f'K{row}'] = value.get('Граница участка дороги', {}).get('количество полос')[0][0]

            for idx, category in enumerate(tuple_cateregory):
                # задаю рамки для всех клеток в диапазоне
                for col in range(1, 12):
                    ws.cell(row = row, column = col).border = self.table_cells_border
                    ws.cell(row = row, column = col).alignment = self.table_cells_aligment
                    ws.cell(row = row, column = col).font = self.table_cells_font
                # объединяю ячейки
                ws.merge_cells(f'E{row}:H{row}')
                ws.merge_cells(f'I{row}:J{row}')
                if category[0] == last_cat[0] and tuple_coating[idx][0] == last_coating[0] and idx != 0:
                    ws[f'A{row}'] = last_cat[-2][0]
                    ws[f'B{row}'] = last_cat[-2][1]
                    ws[f'C{row}'] = category[-1][0]
                    ws[f'D{row}'] = category[-1][1]
                elif category[0] != last_cat[0]:
                    ws[f'C{row}'] = category[-2][0]
                    ws[f'D{row}'] = category[-2][1]
                    row += 1
                    ws[f'A{row}'] = category[-2][0]
                    ws[f'B{row}'] = category[-2][1]
                    ws[f'C{row}'] = category[-1][0]
                    ws[f'D{row}'] = category[-1][1]
                    ws[f'E{row}'] = category[0]
                    ws[f'I{row}'] = tuple_coating[idx][0]
                    ws[f'K{row}'] = value.get('Граница участка дороги', {}).get('количество полос')[idx][0]
                elif tuple_coating[idx][0] != last_coating[0]:
                    ws[f'C{row}'] = tuple_coating[idx][-2][0]
                    ws[f'D{row}'] = tuple_coating[idx][-2][1]
                    row += 1
                    ws[f'A{row}'] = tuple_coating[idx][-2][0]
                    ws[f'B{row}'] = tuple_coating[idx][-2][1]
                    ws[f'C{row}'] = tuple_coating[idx][-1][0]
                    ws[f'D{row}'] = tuple_coating[idx][-1][1]
                    ws[f'E{row}'] = category[0]
                    ws[f'I{row}'] = tuple_coating[idx][0]
                    ws[f'K{row}'] = value.get('Граница участка дороги', {}).get('количество полос')[idx][0]

                last_cat = category
                last_coating = tuple_coating[idx]

    def get_category (self, list_width: list[tuple], value):
        '''
        объединяет и соотносит створы с категориями, количеством полос, видом покрытия
        '''
        res = []
        tuple_cat = tuple((category[0], float(count_line[0]), *tip_travel_clothing) for category,
        count_line, tip_travel_clothing in zip(value.get('Граница участка дороги', {}).get('категория дорог/улиц'),
                                               value.get('Граница участка дороги', {}).get('количество полос'),
                                               value.get('Граница участка дороги', {}).get('тип дорожной одежды')))
        for width in list_width:
            for idx, category in enumerate(tuple_cat):
                if width[-2] <= category[-2]:
                    res.append((*tuple_cat[idx - 1][0:3],))
                    break
                elif width[-2] > tuple_cat[-1][-2]:
                    res.append((*tuple_cat[-1][0:3],))
                    break
                elif width[-2] > category[-2]:
                    continue
        # print(list_cat)
        return res

    def write_roadway_width (self):
        row = 8
        positive_counter = 0
        negative_counter = 0
        length = 0

        list_difference_width = []
        ws = self.wb['Ширина проезжей части']
        for key, value in self.data.items():
            if isinstance(value, str):
                continue
            one_width = value.get('Ширина проезжей части').get('Ширина ПЧ')[0]
            tuple_roadway_width = tuple(
                (float(i[0]), *i[1:]) for i in value.get('Ширина проезжей части').get('Ширина ПЧ')[1:])
            list_category = self.get_category(tuple_roadway_width, value)
            required_width = self.DICT_NORMATIVE_VALUE.get(list_category[0][0])[0] * list_category[0][1]
            # list_required_width.append(roadway_width[0][3] - required_width)
            difference_width = float(one_width[0]) - required_width
            list_difference_width.append(difference_width)
            if len(self.data) > 2:
                # ws.unmerge_cells(f'A{row}:I{row}')
                ws.merge_cells(f'A{row}:I{row}')
                ws[f'A{row}'] = key
                row += 1
            ws[f'A{row}'] = 0  # roadway_width[0][-2][0]  # начало км
            ws[f'B{row}'] = 0  # roadway_width[0][-2][1]  # начало м
            ws[f'C{row}'] = tuple_roadway_width[0][-1][0]  # конец км
            ws[f'D{row}'] = tuple_roadway_width[0][-1][1]  # конец м
            ws[f'E{row}'] = tuple_roadway_width[0][-4]  # протяженность
            ws[f'F{row}'] = one_width[0]  # измеренная
            ws[f'G{row}'] = required_width

            ws[f'H{row}'] = difference_width  # разница
            if abs(difference_width) <= 0.5:
                ws[f'I{row}'] = 'Соответсвует'
                positive_counter += abs(tuple_roadway_width[0][-4])
            else:
                ws[f'I{row}'] = 'Не соответствует'
                negative_counter += abs(tuple_roadway_width[0][-4])
            length = value.get('Ось дороги').get('Начало трассы')[0][2]

            for idx, width in enumerate(tuple_roadway_width):

                for col in range(1, 10):
                    ws.cell(row = row + 1, column = col).border = self.table_cells_border
                    ws.cell(row = row + 1, column = col).alignment = self.table_cells_aligment
                    ws.cell(row = row + 1, column = col).font = self.table_cells_font
                # if idx == 0:
                #     last_width=0

                if idx != len(tuple_roadway_width) - 1:
                    last_width = tuple_roadway_width[idx + 1]
                else:
                    last_width = (width[0], *value.get('Ось дороги').get('Начало трассы')[0])

                    # ws[f'C{row}'] = width[-2][0]  # конец км
                    # ws[f'D{row}'] = width[-2][1]  # конец м
                    # ws[f'E{row}'] = width[-4] - last_width[-4]  # протяженность
                required_width = self.DICT_NORMATIVE_VALUE.get(list_category[idx][0])[0] * list_category[idx][
                    1]  # Требуемая
                difference_width = width[0] - required_width
                list_difference_width.append(difference_width)
                lenght_segment = last_width[-3] - width[-3]
                row += 1
                ws[f'A{row}'] = width[-2][0]
                ws[f'B{row}'] = width[-2][1]
                ws[f'C{row}'] = last_width[-1][0]
                ws[f'D{row}'] = last_width[-1][1]
                ws[f'E{row}'] = last_width[-3] - width[-3]  # протяженность
                ws[f'F{row}'] = width[0]  # измеренная
                ws[f'G{row}'] = required_width  # Требуемая
                ws[f'H{row}'] = difference_width  # разница
                if abs(difference_width) <= 0.5:
                    ws[f'I{row}'] = 'Соответсвует'
                    positive_counter += abs(lenght_segment)
                else:
                    ws[f'I{row}'] = 'Не соответствует'
                    negative_counter += abs(lenght_segment)

        row += 2

        ws[f'I{row}'] = f'Протяженность: {length} м'
        ws[f'I{row}'].alignment = self.total_cells_aligment
        ws[f'I{row + 1}'] = f'Соответствует: {positive_counter} м'
        ws[f'I{row + 1}'].alignment = self.total_cells_aligment
        ws[f'I{row + 2}'] = f'Не соответствует: {negative_counter} м'
        ws[f'I{row + 2}'].alignment = self.total_cells_aligment
        list_difference_width = tuple(list_difference_width)
        # print(list_required_width)
        self.create_bar_diagram(differences = list_difference_width[:-1],
                                title = ("Разница ширины проезжей части от требуемого значения по расстоянию",
                                         'Общая оценка соответствия ширины проезжей части'),
                                calculation_object = [obj for value in self.data.values() if isinstance(value, dict)
                                                       for obj in value.get(
                                         'Ширина проезжей части').get('Ширина ПЧ')][1:],
                                pos_neg_all = (positive_counter, negative_counter, length), page = 'Диаграммы')

    def create_bar_diagram (self, page: str = None, differences = None, title: tuple[str, str] = None, calculation_object = None,
                            pos_neg_all = None, required=None):
        '''
        создание столбчатой диаграммы
        :param page: страница для диаграммы
        :param differences: кортеж с разницей ширины
        :param title: заголовок диаграммы
        :param calculation_object: привязки к объектам
        :param pos_neg_all: кортеж с количеством положительных и отрицательных и суммарно значений
        :param required: требуемое значение для сравнения
        :return:
        '''
        color = FontStyle(color = 'ffffff')
        ws = self.wb[page]
        row = 1
        ws.cell(row = row, column = 2, value = 'Соответствует')
        ws.cell(row = row, column = 2).font = color
        ws.cell(row = row, column = 3, value = 'Не соответствует')
        ws.cell(row = row, column = 3).font = color
        row += 1

        for calc_obj, diff in zip(calculation_object, differences):

            ws.cell(row = row, column = 1, value = calc_obj[-3])
            ws.cell(row = row, column = 1).font = color
            if abs(diff) <= 0.5 or (required is not None and abs(diff) >= required):

                ws.cell(row = row, column = 2, value = diff)
                ws.cell(row = row, column = 2).font = color
                ws.cell(row = row, column = 3, value = 0)
                ws.cell(row = row, column = 3, value = 0).font = color
            else:
                ws.cell(row = row, column = 2, value = 0)
                ws.cell(row = row, column = 2).font = color
                ws.cell(row = row, column = 3, value = diff)
                ws.cell(row = row, column = 3).font = color

            row += 1

        ws.cell(row = row, column = 1, value = pos_neg_all[2])
        ws.cell(row = row, column = 1).font = color
        if abs(differences[-1]) <= 0.5 or (required is not None and abs(diff) >= required):
            ws.cell(row = row, column = 2, value = differences[-1])
            ws.cell(row = row, column = 2).font = color
            ws.cell(row = row, column = 3, value = 0)
            ws.cell(row = row, column = 3, value = 0).font = color
        else:
            ws.cell(row = row, column = 2, value = 0)
            ws.cell(row = row, column = 2).font = color
            ws.cell(row = row, column = 3, value = differences[-1])
            ws.cell(row = row, column = 3).font = color

        row += 1

        ws.cell(row = row, column = 2, value = pos_neg_all[0])
        ws.cell(row = row, column = 2).font = color
        ws.cell(row = row, column = 3, value = pos_neg_all[1])
        ws.cell(row = row, column = 3).font = color
        # ДИАГРАММА №1
        # создаем объект диаграммы
        chart = BarChart(gapWidth = 0, overlap = 100)
        # установим тип - `вертикальные столбцы`
        chart.type = "col"
        chart.grouping = "clustered"
        # установим стиль диаграммы (цветовая схема)
        chart.style = 2
        # заголовок диаграммы
        chart.title = title[0]
        # подпись оси `y`
        chart.y_axis.title = 'Допустимый диапазон'
        # показывать данные на оси
        chart.y_axis.delete = False
        # подпись оси `x`
        chart.x_axis.delete = False
        # выберем 2 столбца с данными для оси `y`
        y = Reference(ws, min_col = 2, max_col = 3, min_row = 1, max_row = row - 1)
        # выберем категорию для оси `x`
        x = Reference(worksheet = ws, min_col = 1, max_col = 1, min_row = 2, max_row = row - 1)
        # добавляем данные в объект диаграммы
        chart.add_data(y, titles_from_data = True)
        # установим метки на объект диаграммы
        chart.set_categories(x)

        # создаем серию для положительных значений
        positive_series = chart.series[0]

        positive_series.graphicalProperties.solidFill = '008000'  # зеленый цвет

        # создаем серию для отрицательных значений
        negative_series = chart.series[1]
        negative_series.graphicalProperties.solidFill = 'FF0000'  # красный цвет

        # добавляем легенду
        legend = chart.legend
        legend.position = 'b'
        legend.include_in_layout = False

        chart.width = 19
        chart.height = 10
        self.set_chart_text_style(chart)

        # добавляем диаграмму на лист
        ws.add_chart(chart, 'A1')
        self.create_pie_chart(ws, row, title = title[1])

    def create_pie_chart (self, ws, row, title = None):
        # круговая диаграмма
        chart_pie = PieChart()
        chart_pie.type = "bar"
        chart_pie.grouping = "stacked"
        chart_pie.style = 2

        chart_pie.title = title
        chart_pie.width = 19
        chart_pie.height = 10
        labels = Reference(ws, min_row = 1, max_row = 1, min_col = 2, max_col = 3)
        data = Reference(ws, min_row = row, max_row = row, min_col = 2, max_col = 3)
        chart_pie.add_data(data, from_rows = True)
        chart_pie.set_categories(labels)
        chart_pie.dataLabels = DataLabelList()
        # Show the percent on the labels.
        chart_pie.dataLabels.showPercent = True
        chart_pie.dataLabels.position = 'ctr'

        series = chart_pie.series[0]
        positive_dp = DataPoint(idx = 0)
        negative_dp = DataPoint(idx = 1)
        negative_dp.graphicalProperties.solidFill = 'FF0000'
        positive_dp.graphicalProperties.solidFill = '008000'
        series.data_points = [positive_dp, negative_dp]
        series.marker.symbol = "circle"

        legend = chart_pie.legend
        legend.position = 'b'
        legend.include_in_layout = False

        ws.add_chart(chart_pie, 'A20')
        self.set_chart_text_style(chart_pie)

    def write_shoulder_width (self):
        ws = self.wb['Ширина обочин']
        row = 9
        positive_counter = 0
        negative_counter = 0
        length = 0
        list_difference_width = []
        widths_all = []
        for key, value in self.data.items():
            if isinstance(value, str):
                continue
            w = value.get('Ширина обочин').get('Ширина обочины')
            widths = []
            direction = value.get('Ширина обочин').get('Направление')
            for idx, width in enumerate(w):

                #ic(direction_shoulder_width[idx])
                if idx % 2 == 0:
                    if direction[idx][0] == 'Прямое':
                        widths.append((float(width[0]), float(w[idx+1][0]), *width[1:]))
                    else:
                        widths.append((float(w[idx+1][0]), float(width[0]), *width[1:]))
            one_width = widths.pop(0)
            list_category = self.get_category(widths, value)
            required_width = self.DICT_NORMATIVE_VALUE.get(list_category[0][0])[1]
            difference_width = float(one_width[0]) - required_width
            list_difference_width.append(difference_width)

            if len(self.data) > 2:
                # ws.unmerge_cells(f'A{row}:I{row}')
                ws.merge_cells(f'A{row}:K{row}')
                ws[f'A{row}'] = key
                row += 1
            ws[f'A{row}'] = 0  # roadway_width[0][-2][0]  # начало км
            ws[f'B{row}'] = 0  # roadway_width[0][-2][1]  # начало м
            ws[f'C{row}'] = widths[0][-1][0]  # конец км
            ws[f'D{row}'] = widths[0][-1][1]  # конец м
            ws[f'E{row}'] = widths[0][-4]  # протяженность
            ws[f'F{row}'] = one_width[0]  # измеренная слева
            ws[f'G{row}'] = one_width[1] # измеренная справо
            ws[f'H{row}'] = min(one_width[:2]) #наименьшее
            ws[f'I{row}'] = required_width #требуемое
            ws[f'J{row}'] = difference_width  # разница
            if abs(difference_width) <= 0.5:
                ws[f'K{row}'] = 'Соответсвует'
                positive_counter += abs(widths[0][-4])
            else:
                ws[f'K{row}'] = 'Не соответствует'
                negative_counter += abs(widths[0][-4])
            length = value.get('Ось дороги').get('Начало трассы')[0][2]
            for idx, width in enumerate(widths):

                for col in range(1, 12):
                    ws.cell(row = row + 1, column = col).border = self.table_cells_border
                    ws.cell(row = row + 1, column = col).alignment = self.table_cells_aligment
                    ws.cell(row = row + 1, column = col).font = self.table_cells_font

                if idx != len(widths) - 1:
                    last_width = widths[idx + 1]
                else:
                    last_width = (width[0], *value.get('Ось дороги').get('Начало трассы')[0])

                required_width = self.DICT_NORMATIVE_VALUE.get(list_category[idx][0])[1]  #list_category[idx][1]  # Требуемая
                difference_width = width[0] - required_width
                list_difference_width.append(difference_width)
                lenght_segment = last_width[-3] - width[-3]
                row += 1
                ws[f'A{row}'] = width[-2][0]
                ws[f'B{row}'] = width[-2][1]
                ws[f'C{row}'] = last_width[-1][0]
                ws[f'D{row}'] = last_width[-1][1]
                ws[f'E{row}'] = lenght_segment
                ws[f'F{row}'] = width[0]  # измеренная слева
                ws[f'G{row}'] = width[1] # измеренная справо
                ws[f'H{row}'] = min(width[:2]) #наименьшее
                ws[f'I{row}'] = required_width #требуемое
                ws[f'J{row}'] = difference_width  # разница

                if abs(difference_width) <= 0.5:
                    ws[f'K{row}'] = 'Соответсвует'
                    positive_counter += abs(lenght_segment)
                else:
                    ws[f'K{row}'] = 'Не соответствует'
                    negative_counter += abs(lenght_segment)
            widths_all.extend(widths)
        row += 2
        ws[f'K{row}'] = f'Протяженность: {length} м'
        ws[f'K{row}'].alignment = self.total_cells_aligment
        ws[f'K{row + 1}'] = f'Соответствует: {positive_counter} м'
        ws[f'K{row + 1}'].alignment = self.total_cells_aligment
        ws[f'K{row + 2}'] = f'Не соответствует: {negative_counter} м'
        ws[f'K{row + 2}'].alignment = self.total_cells_aligment
        self.create_bar_diagram(differences = list_difference_width[:-1],
                                title = ("Разница ширины обочины от требуемого значения по расстоянию",
                                         "Общая оценка соответствия ширины обочины"),
                                calculation_object = widths_all[1:],
                                pos_neg_all = (positive_counter, negative_counter, length), page = 'Диаграммы 1')

    def write_curves(self):
        ws = self.wb['Радиусы кривых в плане']
        row = 7
        positive_counter = 0
        negative_counter = 0
        length = 0
        curves_all_values = []
        required_min_curve = 0
        for key, value in self.data.items():
            if isinstance(value, str):
                continue
            if len(self.data) > 2:
                # ws.unmerge_cells(f'A{row}:I{row}')
                ws.merge_cells(f'A{row}:K{row}')
                ws[f'A{row}'] = key
                row += 1
            curves = [curve for curve in value.get('Кривая', {}).get('R', []) if '.' in curve[0] and float(curve[0])!= 0]
            list_category = self.get_category(curves, value)

            for idx,curve in enumerate(curves):
                required_min_curve = self.DICT_NORMATIVE_VALUE.get(list_category[idx][0])[3]
                required_speed = self.DICT_NORMATIVE_VALUE.get(list_category[idx][0])[2]
                for col in range(1, 9):
                    ws.cell(row = row + 1, column = col).border = self.table_cells_border
                    ws.cell(row = row + 1, column = col).alignment = self.table_cells_aligment
                    ws.cell(row = row + 1, column = col).font = self.table_cells_font
                ws[f'A{row}'] = curve[-2][0]  # начало км
                ws[f'B{row}'] = curve[-2][1]  # начало м
                ws[f'C{row}'] = curve[-1][0]  # конец км
                ws[f'D{row}'] = curve[-1][1]  # конец м
                ws[f'E{row}'] = required_speed  # расчетная скорость
                ws[f'F{row}'] = round(float(curve[0]),2)  # измеренный радиус
                ws[f'G{row}'] = required_min_curve  # требуемый радиус
                curves_all_values.append(curve)
                if abs(float(curve[0])) >= required_min_curve:
                    ws[f'H{row}'] = 'Соответсвует'
                    positive_counter += 1
                else:
                    ws[f'H{row}'] = 'Не соответствует'
                    negative_counter += 1
                row+=1
                length = value.get('Ось дороги').get('Начало трассы')[0][2]
        # row += 2
        # ws[f'H{row}'] = f'Протяженность: {length} м'
        # ws[f'H{row}'].alignment = self.total_cells_aligment
        # ws[f'H{row + 1}'] = f'Соответствует: {positive_counter} м'
        # ws[f'H{row + 1}'].alignment = self.total_cells_aligment
        # ws[f'H{row + 2}'] = f'Не соответствует: {negative_counter} м'
        # ws[f'H{row + 2}'].alignment = self.total_cells_aligment
        calc_obj = [(*curve[:-3],'\n'.join(self.change_start_and_end_obj(*curve[-2:])),*curve[-2:])for curve in curves_all_values ]


        self.create_bar_diagram(differences = [float(curve[0]) for curve in curves_all_values],
                                title = ("Оценка соответствиявеличины радиусов кривых в плане по расстоянию",
                                         "Общая оценка соответствия радиусов кривых в плане"),
                                calculation_object = calc_obj,
                                pos_neg_all = (positive_counter, negative_counter, length), page = 'Диаграммы 2',
                                required = required_min_curve)

class WriterApplication(WriterExcel):
    def __init__ (self, data: dict = None, path: str = None, data_interface: dict = None):
        super().__init__(data = data, path_template_excel = path_template_excel_application, path = path,
                         data_interface = data_interface)
        self.tip_doc = 'Приложение'


class WriterApplicationCityTP(WriterApplication):
    def __init__ (self, data: dict = None, path = None, data_interface = None):
        super().__init__(data = data, path = path, data_interface = data_interface)

        self.cells_font_result = FontStyle(name = 'Times New Roman', size = 12, bold = True)
        self.table_cells_font = FontStyle(name = 'Times New Roman', size = 12)
        thin = Side(border_style = "thin", color = "000000")
        # thick = Side(border_style = "thick", color = "000000")
        self.table_cells_border = Border(left = thin, right = thin, top = thin, bottom = thin, )
        # self.table_cells_aligment = Alignment(horizontal = 'center', vertical = 'center')
        self.table_cells_aligment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
        self.cells_result = Alignment(horizontal = 'right', )
        self.cells_result_value = Alignment(horizontal = 'left')
        print("#############################\nНачал формировать ведомости!\n#############################\n")
        self.run()
        print('сохранение')
        self.save_file()

    def run (self):
        print('write_roadway')
        self.write_roadway()
        print('write_separator_strip')
        self.write_separator_strip()
        print('write_reinforced_shoulders')
        self.write_reinforced_shoulders()
        print('write_exit_road')
        self.write_exit_road()
        print('write_other_territories')
        self.write_other_territories()
        print('write_sidewalk')
        self.write_sidewalk()
        print('write_border')
        self.write_border()
        print('write_luke')
        self.write_luke()
        print('write_other_engineering_structures')
        self.write_other_engineering_structures()
        print('write_bus_stop')
        self.write_bus_stop()
        print('write_lighting_poles')
        self.write_lighting_poles()
        print('write_maf')
        self.write_maf()
        print('write_signs')
        self.write_signs()
        print('write_fence')
        self.write_fence()
        print('write_traffic_light')
        self.write_traffic_light()
        print('write_communications')
        self.write_communications()
        print('write_pipes')
        self.write_pipes()
        print('write_bridge')
        self.write_bridge()
        print('write_turns')
        self.write_turns()
        print('write_gazon')
        self.write_gazon()

    def write_roadway (self):
        """заполнеие табилц проезжая часть"""
        ws = self.wb['ПЧ']
        row = 9
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue

            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:K{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1

            for idx, value in enumerate(v1.get('Проезжая часть', {}).get('Назначение', [])):
                if value[0] == 'основные полосы движения':
                    counter += 1
                    ws[f'B{row}'] = counter
                    ws[f'B{row}'].border = self.table_cells_border
                    ws[f'B{row}'].alignment = self.table_cells_aligment
                    ws[f'B{row}'].font = self.table_cells_font

                    ws[f'C{row}'] = f'{value[-2][0]} + 000' if value[-2][1] == 0 else f'{value[-2][0]} + {value[-2][1]}'
                    ws[f'C{row}'].border = self.table_cells_border
                    ws[f'C{row}'].alignment = self.table_cells_aligment
                    ws[f'C{row}'].font = self.table_cells_font

                    ws[f'D{row}'] = f'{value[-1][0]} + 000' if value[-1][1] == 0 else f'{value[-1][0]} + {value[-1][1]}'
                    ws[f'D{row}'].border = self.table_cells_border
                    ws[f'D{row}'].alignment = self.table_cells_aligment
                    ws[f'D{row}'].font = self.table_cells_font

                    ws[f'E{row}'] = value[-1][1] - value[-2][1]
                    ws[f'E{row}'].border = self.table_cells_border
                    ws[f'E{row}'].alignment = self.table_cells_aligment
                    ws[f'E{row}'].font = self.table_cells_font

                    ws[f'F{row}'] = 'оба' if v1.get('Проезжая часть', {}).get('Расположение', [])[idx][
                                                 0] == 'По оси' else \
                        v1.get('Проезжая часть', {}).get('Расположение', [])[idx][0]
                    ws[f'F{row}'].border = self.table_cells_border
                    ws[f'F{row}'].alignment = self.table_cells_aligment
                    ws[f'F{row}'].font = self.table_cells_font

                    ws[f'G{row}'] = v1.get('Проезжая часть', {}).get('Тип покрытия', [])[idx][0]
                    ws[f'G{row}'].border = self.table_cells_border
                    ws[f'G{row}'].alignment = self.table_cells_aligment
                    ws[f'G{row}'].font = self.table_cells_font

                    ws[f'H{row}'] = value[2]  # square
                    ws[f'H{row}'].border = self.table_cells_border
                    ws[f'H{row}'].alignment = self.table_cells_aligment
                    ws[f'H{row}'].font = self.table_cells_font

                    ws[f'I{row}'].border = self.table_cells_border
                    ws[f'I{row}'].alignment = self.table_cells_aligment
                    ws[f'I{row}'].font = self.table_cells_font

                    ws[f'J{row}'].border = self.table_cells_border
                    ws[f'J{row}'].alignment = self.table_cells_aligment
                    ws[f'J{row}'].font = self.table_cells_font

                    ws[f'K{row}'].border = self.table_cells_border
                    ws[f'K{row}'].alignment = self.table_cells_aligment
                    ws[f'K{row}'].font = self.table_cells_font

                    row += 1

            counter_sum += counter
        ws[f'J{row + 2}'] = 'Итого протяженность (м):'
        ws[f'J{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'J{row + 2}'].font = self.cells_font_result

        ws[f'K{row + 2}'] = f'=SUM(E9:E{row})'
        ws[f'K{row + 2}'].alignment = self.cells_result_value  # выравнивание по левому краю
        ws[f'K{row + 2}'].font = self.table_cells_font

        ws[f'J{row + 3}'] = 'Итого площадь (м²):'
        ws[f'J{row + 3}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'J{row + 3}'].font = self.cells_font_result

        ws[f'K{row + 3}'] = f'=SUM(H9:H{row})'
        ws[f'K{row + 3}'].alignment = self.cells_result_value  # выравнивание по левому краю
        ws[f'K{row + 3}'].font = self.table_cells_font
        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_separator_strip (self):

        """
        заполенение таблицы разделительные полосы
        :return:
        """
        counter_sum = 0
        row = 9
        ws = self.wb['разделительная полоса']

        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:L{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1
            for idx, value in enumerate(v1.get('Разделительная полоса', {}).get('Расположение', [])):
                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = value[-2][1]
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-1][1]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[-1][1] - value[-2][1]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = value[0]
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = v1.get('Разделительная полоса', {}).get('Тип покрытия', [])[idx][0] if idx < len(
                    v1.get('Разделительная полоса', {}).get('Тип покрытия', [])) else ''
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'] = round(value[2] / (value[-1][1] - value[-2][1]), 2)
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = value[2]
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = ''
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                ws[f'K{row}'] = ''
                ws[f'K{row}'].border = self.table_cells_border
                ws[f'K{row}'].alignment = self.table_cells_aligment
                ws[f'K{row}'].font = self.table_cells_font

                ws[f'L{row}'] = ''
                ws[f'L{row}'].border = self.table_cells_border
                ws[f'L{row}'].alignment = self.table_cells_aligment
                ws[f'L{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
        ws[f'K{row + 2}'] = 'Итого протяженность (м):'
        ws[f'K{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'K{row + 2}'].font = self.cells_font_result

        ws[f'L{row + 2}'] = f'=SUM(E9:E{row})'
        ws[f'L{row + 2}'].alignment = self.cells_result_value  # выравнивание по левому краю
        ws[f'L{row + 2}'].font = self.table_cells_font

        ws[f'K{row + 3}'] = 'Итого площадь (м2): '
        ws[f'K{row + 3}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'K{row + 3}'].font = self.cells_font_result

        ws[f'L{row + 3}'] = f'=SUM(I9:I{row})'
        ws[f'L{row + 3}'].alignment = self.cells_result_value  # выравнивание по левому краю
        ws[f'L{row + 3}'].font = self.table_cells_font
        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_reinforced_shoulders (self):
        """
        заполенение таблицы наличие укрепленных обочин
        :return:
        """

        row = 8
        ws = self.wb['укреп. обочины']
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'D{row}:K{row}')
                ws[f'D{row}'] = k1
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font
                row += 1
            for idx, value in enumerate(v1.get('Укрепленная часть обочины', {}).get('Расположение', [])):
                counter += 1
                ws[f'D{row}'] = counter
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[-2][1]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = value[-1][1]
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = value[0]
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'] = value[-1][1] - value[-2][1]
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = v1.get('Укрепленная часть обочины').get('Тип покрытия', [])[idx][0] if idx < len(
                    v1.get('Укрепленная часть обочины').get('Тип покрытия', [])) else ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = value[2]
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                ws[f'K{row}'] = ''
                ws[f'K{row}'].border = self.table_cells_border
                ws[f'K{row}'].alignment = self.table_cells_aligment
                ws[f'K{row}'].font = self.table_cells_font

                row += 1
            counter_sum += counter
        ws[f'J{row + 2}'] = 'Итого протяженность (м):'
        ws[f'J{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'J{row + 2}'].font = self.cells_font_result

        ws[f'K{row + 2}'] = f'=SUM(H8:H{row})'
        ws[f'K{row + 2}'].alignment = self.cells_result_value  # выравнивание по левому краю
        ws[f'K{row + 2}'].font = self.table_cells_font
        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_exit_road (self):

        """
        заполенение таблицы съездов
        :return:
        """
        counter_sum = 0
        row = 8
        ws = self.wb['съезды']

        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                for col in range(3, 12):
                    ws.cell(row = row, column = col).border = self.table_cells_border
                ws.merge_cells(f'C{row}:K{row}')
                ws[f'C{row}'] = k1
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font
                row += 1
            for idx, value in enumerate(v1.get('Съезд', {}).get('Расположение', [])):
                counter += 1
                ws[f'C{row}'] = counter
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-2][1]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[-1][1]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = 'чет.' if value[6] > 0 else 'нечет.'
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = 'Съезд'
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'] = v1.get('Съезд').get('Тип покрытия', [])[idx][0] if idx < len(
                    v1.get('Съезд').get('Тип покрытия', [])) else ''
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = value[2]
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = v1.get('Съезд').get('Назначение съезда', [])[idx][0] if idx < len(
                    v1.get('Съезд').get('Назначение съезда', [])) else ''
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                if len(ws[f'J{row}'].value) > 15:
                    ws.row_dimensions[row].height = 31.48

                ws[f'K{row}'] = ''
                ws[f'K{row}'].border = self.table_cells_border
                ws[f'K{row}'].alignment = self.table_cells_aligment
                ws[f'K{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
        ws[f'K{row + 2}'] = f'Итого (шт):{counter_sum}'
        ws[f'K{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'K{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_other_territories (self):

        """
        заполенение таблицы съездов
        :return:
        """
        counter_sum = 0
        row = 8
        ws = self.wb['прочие территории']

        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'C{row}:J{row}')
                ws[f'C{row}'] = k1
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font
                row += 1
            for idx, value in enumerate(v1.get('Проезжая часть', {}).get('Назначение', [])):
                if value[0] in ['площадка отдыха', 'автостоянка', 'парковка', 'отстоно-разворотная площадка',
                                'трамвайное полотно']:
                    counter += 1
                    ws[f'C{row}'] = counter
                    ws[f'C{row}'].border = self.table_cells_border
                    ws[f'C{row}'].alignment = self.table_cells_aligment
                    ws[f'C{row}'].font = self.table_cells_font

                    ws[f'D{row}'] = 'чет.' if value[6] > 0 else 'нечет.'
                    ws[f'D{row}'].border = self.table_cells_border
                    ws[f'D{row}'].alignment = self.table_cells_aligment
                    ws[f'D{row}'].font = self.table_cells_font

                    ws[f'E{row}'] = value[-2][1]
                    ws[f'E{row}'].border = self.table_cells_border
                    ws[f'E{row}'].alignment = self.table_cells_aligment
                    ws[f'E{row}'].font = self.table_cells_font

                    ws[f'F{row}'] = value[-1][1]
                    ws[f'F{row}'].border = self.table_cells_border
                    ws[f'F{row}'].alignment = self.table_cells_aligment
                    ws[f'F{row}'].font = self.table_cells_font

                    ws[f'G{row}'] = value[0]
                    ws[f'G{row}'].border = self.table_cells_border
                    ws[f'G{row}'].alignment = self.table_cells_aligment
                    ws[f'G{row}'].font = self.table_cells_font

                    ws[f'H{row}'] = value[2]
                    ws[f'H{row}'].border = self.table_cells_border
                    ws[f'H{row}'].alignment = self.table_cells_aligment
                    ws[f'H{row}'].font = self.table_cells_font

                    ws[f'I{row}'] = ''
                    ws[f'I{row}'].border = self.table_cells_border
                    ws[f'I{row}'].alignment = self.table_cells_aligment
                    ws[f'I{row}'].font = self.table_cells_font

                    ws[f'J{row}'] = ''
                    ws[f'J{row}'].border = self.table_cells_border
                    ws[f'J{row}'].alignment = self.table_cells_aligment
                    ws[f'J{row}'].font = self.table_cells_font

                    row += 1

            counter_sum += counter
        ws[f'J{row + 2}'] = f'Итого (шт.): {counter_sum}'
        ws[f'J{row + 2}'].alignment = self.cells_result
        ws[f'J{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_sidewalk (self):
        """
             заполенение таблицы тротуары
             :return:
             """
        row = 8
        ws = self.wb['тротуары']
        counter_sum = 0
        lenght_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'D{row}:L{row}')
                ws[f'D{row}'] = k1
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                row += 1

            for idx, value in enumerate(v1.get('Тротуар', {}).get('Расположение', [])):
                counter += 1
                ws[f'D{row}'] = counter
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = "чет." if value[6] > 0 else 'нечет.'
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = value[-2][1]
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = value[-1][1]
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                lenght_sum += abs(value[-1][1] - value[-2][1])
                ws[f'H{row}'] = value[-1][1] - value[-2][1] if value[-1][1] - value[-2][1] != 0 else ''
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = round(value[2] / (value[-1][1] - value[-2][1]), 2) if value[-1][1] - value[-2][
                    1] != 0 else ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = value[2]
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                ws[f'K{row}'] = v1.get('Тротуар').get('Материал покрытия', [])[idx][0] if idx < len(
                    v1.get('Тротуар').get('Материал покрытия', [])) else ''
                ws[f'K{row}'].border = self.table_cells_border
                ws[f'K{row}'].alignment = self.table_cells_aligment
                ws[f'K{row}'].font = self.table_cells_font

                ws[f'L{row}'] = ''
                ws[f'L{row}'].border = self.table_cells_border
                ws[f'L{row}'].alignment = self.table_cells_aligment
                ws[f'L{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
        lenght_sum = round(lenght_sum, 2)
        ws[f'L{row + 2}'] = f'Итого (п.м.): {lenght_sum}'
        ws[f'L{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'L{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_border (self):
        """
        заполенение таблицы бордюры
        Должны были добавить расположение в бордюры
        :return:
        """

        row = 8
        ws = self.wb['бордюр']
        counter_sum = 0
        lenght_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:I{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1
            for idx, value in enumerate(v1.get('Бордюр', {}).get('Назначение', [])):
                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = 'чет.' if value[6] > 0 else 'нечет.'
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-2][1]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[-1][1]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                lenght_sum += value[1]
                ws[f'F{row}'] = value[1]
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = value[0]
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'] = v1.get('Бордюр', {}).get('Марка')[idx][0] if idx < len(
                    v1.get('Бордюр', {}).get('Марка', [])) else ''
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
        lenght_sum = round(lenght_sum, 2)
        ws[f'I{row + 2}'] = f'Итого (п.м):{lenght_sum}'
        ws[f'I{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'I{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_luke (self):

        """
        заполенение таблицы люки
        :return:
        """

        row = 8
        ws = self.wb['люки']
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:K{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1

            tips_luke = ('Люк смотрового колодца', 'Решетка дождеприемного колодца')
            luks = []
            for tip in tips_luke:
                for value in v1.get(tip, {}).get('Расположение', []):
                    luks.append((tip, *value))
            luks.sort(key = lambda x: x[-2])

            for value in luks:
                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = 'чет.' if value[6] > 0 else 'нечет.'
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-2][1]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = abs(round(value[6], 1))
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = '+' if value[1] == 'ПЧ' else ''
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = '+' if value[1] == 'Тротуар' else ''
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'] = '+' if value[1] == 'Газон' else ''
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = 'смотровой' if value[0] == 'Люк смотрового колодца' else 'ливневый'
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = ''
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                ws[f'K{row}'] = ''
                ws[f'K{row}'].border = self.table_cells_border
                ws[f'K{row}'].alignment = self.table_cells_aligment
                ws[f'K{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
        ws[f'K{row + 2}'] = f'Итого (шт.):{counter_sum}'
        ws[f'K{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'K{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_other_engineering_structures (self):

        """
        заполенение таблицы прочие инженерные сооружения
        :return:
        """
        # продольный лоток перезд подпорная стенка подземный пешеходный переход
        row = 8
        ws = self.wb['прочие инж.сооруж.']
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:I{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1
            diсt_obj = {'Железнодорожный переезд': v1.get('Железнодорожный переезд', {}).get('Материал настила', []),
                        'Продольный лоток': v1.get('Продольный лоток', {}).get('Материал', []),
                        'Подпорная стенка': v1.get('Подпорная стенка', {}).get('Материал', []),
                        'Пешеходный переход подземный': v1.get('Пешеходный переход подземный', {}).get('Техническое '
                                                                                                       'состояние', [])}

            for key, value_list in diсt_obj.items():

                for value in value_list:
                    counter += 1
                    ws[f'B{row}'] = counter
                    ws[f'B{row}'].border = self.table_cells_border
                    ws[f'B{row}'].alignment = self.table_cells_aligment
                    ws[f'B{row}'].font = self.table_cells_font

                    ws[f'C{row}'] = 'чет.' if value[6] > 0 else 'нечет.'
                    ws[f'C{row}'].border = self.table_cells_border
                    ws[f'C{row}'].alignment = self.table_cells_aligment
                    ws[f'C{row}'].font = self.table_cells_font

                    ws[f'D{row}'] = value[-2][1]
                    ws[f'D{row}'].border = self.table_cells_border
                    ws[f'D{row}'].alignment = self.table_cells_aligment
                    ws[f'D{row}'].font = self.table_cells_font

                    ws[f'E{row}'] = value[-1][1]
                    ws[f'E{row}'].border = self.table_cells_border
                    ws[f'E{row}'].alignment = self.table_cells_aligment
                    ws[f'E{row}'].font = self.table_cells_font

                    ws[f'F{row}'] = key
                    ws[f'F{row}'].border = self.table_cells_border
                    ws[f'F{row}'].alignment = self.table_cells_aligment
                    ws[f'F{row}'].font = self.table_cells_font

                    ws[f'G{row}'] = value[0]
                    ws[f'G{row}'].border = self.table_cells_border
                    ws[f'G{row}'].alignment = self.table_cells_aligment
                    ws[f'G{row}'].font = self.table_cells_font

                    ws[f'H{row}'] = value[2]
                    ws[f'H{row}'].border = self.table_cells_border
                    ws[f'H{row}'].alignment = self.table_cells_aligment
                    ws[f'H{row}'].font = self.table_cells_font

                    ws[f'I{row}'] = ''
                    ws[f'I{row}'].border = self.table_cells_border
                    ws[f'I{row}'].alignment = self.table_cells_aligment
                    ws[f'I{row}'].font = self.table_cells_font

                    row += 1

            counter_sum += counter
        ws[f'I{row + 2}'] = f'Итого (шт.):{counter_sum}'
        ws[f'I{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'I{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_bus_stop (self):

        """
        нет объектра кромка. как посчитать растояние от кромки до павильона?
        заполенение таблицы остановка
        :return:
        """

        row = 8
        ws = self.wb['остановки']
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:L{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1
            bus_stop_list = v1.get('Остановка', {}).get('№ карточки', [])

            for idx, value in enumerate(bus_stop_list):
                # фильтрую список посадочных площадок и заездные карманы по номеру карточки
                square_landing_pad = list(
                    filter(lambda x: x[0] == v1.get('Остановка', {}).get('№ карточки', [])[idx][0],
                           v1.get('Посадочная площадка', {}).get('№ карточки', [])))

                square_pocket = list(filter(lambda x: x[0] == v1.get('Остановка', {}).get('№ карточки', [])[idx][0],
                                            v1.get('Заездной карман', {}).get('№ карточки', [])))

                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = 'чет.' if value[6] > 0 else 'нечет.'
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-2][1]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[-1][1]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = v1.get('Остановка', {}).get('Название остановки', [])[idx][0] if idx < len(
                    v1.get('Остановка', {}).get('Название остановки', [])) else ''

                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font
                if len(ws[f'F{row}'].value) > 15:
                    ws.row_dimensions[row].height = 31.7
                sum_pavilion = sum(
                    1 for i in v1.get('Павильон остановки', {}).get('№ карточки', []) if i[0] == value[0])
                ws[f'G{row}'] = sum_pavilion if sum_pavilion != 0 else '-'
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                list_distance_to_pavilion = v1.get('Остановка', {}).get(
                    'Расстояние от кромки проезжей части до павильона', [])
                distance_to_pavilion = [i[0] for i in list_distance_to_pavilion if
                                        i[-2] == value[-2] and i[-1] == value[-1]]
                ws[f'H{row}'] = distance_to_pavilion[0] if distance_to_pavilion else '-'
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = square_landing_pad[0][2] if square_landing_pad else ''
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                ws[f'K{row}'] = square_pocket[0][2] if square_pocket else ''
                ws[f'K{row}'].border = self.table_cells_border
                ws[f'K{row}'].alignment = self.table_cells_aligment
                ws[f'K{row}'].font = self.table_cells_font

                ws[f'L{row}'].border = self.table_cells_border
                ws[f'L{row}'].alignment = self.table_cells_aligment
                ws[f'L{row}'].font = self.table_cells_font
                row += 1

            counter_sum += counter
        ws[f'L{row + 2}'] = f'Итого (шт.):{counter_sum}'
        ws[f'L{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'L{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_lighting_poles (self):
        """
        заполенение таблицы Опоры освещения и контактной сети

        :return:
        """

        row = 8
        ws = self.wb['освещение']
        counter_sum = 0  # количество записей
        count_sum = 0  # количество штук
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:I{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1
            for idx, value in enumerate(v1.get('Опоры освещения и контактной сети', {}).get('Материал опоры', [])):
                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = 'чет.' if value[6] > 0 else 'нечет.'
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-2][1]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[-1][1]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = 'Опоры освещения и контактной сети'
                if len(ws[f'F{row}'].value) > 15:
                    ws.row_dimensions[row].height = 31.7
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = v1.get('Опоры освещения и контактной сети', {}).get('Материал опоры')[idx][
                    0] if idx < len(
                    v1.get('Опоры освещения и контактной сети', {}).get('Материал опоры')) else ''

                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'] = v1.get('Опоры освещения и контактной сети', {}).get('Материал опоры')[idx][
                    4] if idx < len(
                    v1.get('Опоры освещения и контактной сети', {}).get('Материал опоры')) else ''
                count_sum += int(ws[f'H{row}'].value)
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
        ws[f'I{row + 2}'] = f'Итого (шт.):{count_sum}'
        ws[f'I{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'I{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_maf (self):
        """
        заполенение таблицы малых архитектурных форм

        :return:
        """

        row = 8
        ws = self.wb['МАФ']
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0

            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'C{row}:J{row}')
                ws[f'C{row}'] = k1
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font
                row += 1

            for idx, value in enumerate(v1.get('Малая архитектурная форма', {}).get('Вид МАФ', [])):
                counter += 1
                ws[f'C{row}'] = counter
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = 'чет.' if value[6] > 0 else 'нечет.'
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[-2][1]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = value[-1][1]
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = value[0]
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'] = ''
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = ''
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
        ws[f'J{row + 2}'] = f'Итого (шт.):{counter_sum}'
        ws[f'J{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'J{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_signs (self):
        """
        заполение таблицы знаков
        :return:
        """
        row = 8
        ws = self.wb['знаки']
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            list_sign = []
            if isinstance(v1, str):
                continue

            for k2, v2 in v1.items():
                if k2[0].isdigit():
                    for idx, value in enumerate(v2.get('Статус', [])):
                        if value[0] == 'факт':
                            try:
                                list_sign.append((*k2.split(" ", 1), *v2.get('Способ установки')[idx]))
                            except Exception as e:
                                ic(e, k1, k2, v2, value)
                                self.errors.write(
                                    f'{self.tip_doc} Таблица знаки На участке "{k1}" направление {v2["Направление"][idx][0]}'
                                    f' Знак {k2}{value[-1]} отсутствует способ установки')

            list_sign.sort(key = lambda x: x[-2])

            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:J{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1
            for idx, sign in enumerate(list_sign):

                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = 'чет.' if sign[8] > 0 else 'нечет.'
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = sign[-4]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = sign[1]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font
                if len(ws[f'E{row}'].value) > 21:
                    ws.row_dimensions[row].height = 31.7

                ws[f'F{row}'] = ''
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = sign[0]
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'] = sign[2]
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = ''
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment

                row += 1
            counter_sum += counter
        ws[f'J{row + 2}'] = f'Итого (шт.):{counter_sum}'
        ws[f'J{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'J{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_fence (self):
        '''
        заполение таблицы ограждений
        :return:
        '''
        row = 8
        ws = self.wb['ограждения']
        counter_sum = 0
        counter_sum_sign_column = 0
        counter_sum_fence = 0
        check_sign_column = False
        for k1, v1 in self.data.items():
            if isinstance(v1, str):
                continue
            counter = 0

            list_fence = []
            tuple_emumerate_fence = ('Нестандартное ограждение', 'Пешеходное ограждение', 'Тросовое ограждение',
                                     'Типа Нью-Джерси', 'Металическое барьерное ограждение', 'Парапетное ограждение')
            for fence in tuple_emumerate_fence:
                for value in v1.get(fence, {}).get('Статус', []):
                    list_fence.append((fence, *value))

            list_fence.sort(key = lambda x: x[-2])
            ic(list_fence)
            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:I{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1
            for value in list_fence:
                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = 'чет.' if value[7] > 0 else 'нечет.'
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-2][1]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[-1][1]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = value[0]
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                if value[0] == "Сигнальные столбики":
                    ws[f'G{row}'] = 'штук'
                    ws[f'H{row}'] = value[5]
                    check_sign_column = True
                    counter_sum_sign_column += value[5]

                else:
                    ws[f'G{row}'] = 'п.м.'
                    ws[f'H{row}'] = value[2]
                    counter_sum_fence += value[2]
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                row += 1
            counter_sum += counter
            counter_sum_fence = round(counter_sum_fence, 2)
        if check_sign_column:
            ws[f'H{row + 2}'] = f'Итого сигнальных столбиков (шт.):{counter_sum_sign_column}'
            ws[f'H{row + 2}'].alignment = self.cells_result
            ws[f'H{row + 2}'].font = self.cells_font_result

            ws[f'H{row + 3}'] = f'Итого ограждения (п.м):{counter_sum_fence}'
            ws[f'H{row + 3}'].alignment = self.cells_result  # выравнивание по правому краю
            ws[f'H{row + 3}'].font = self.cells_font_result

        else:
            ws[f'H{row + 2}'] = f'Итого ограждения (п.м):{counter_sum_fence}'
            ws[f'H{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
            ws[f'H{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_traffic_light (self):

        """
         заполенение таблицы светофоры

         :return:
         """

        row = 8
        ws = self.wb['светофоры']
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:K{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1

            for idx, value in enumerate(v1.get('Светофор', {}).get('Тип установки', [])):
                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = 'чет.' if value[6] > 0 else 'нечет.'
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-2][1]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[0]
                if len(ws[f'E{row}'].value) > 15:
                    ws.row_dimensions[row].height = 31.7
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = ''
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                if v1.get('Светофор', {}).get('Тип', [])[idx][0].lower() == "транспортный":
                    ws[f'G{row}'] = '+'
                elif v1.get('Светофор', {}).get('Тип', [])[idx][0].lower() == "пешеходный":
                    ws[f'H{row}'] = '+'

                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = ''
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                ws[f'K{row}'] = ''
                ws[f'K{row}'].border = self.table_cells_border
                ws[f'K{row}'].alignment = self.table_cells_aligment
                ws[f'K{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
        ws[f'K{row + 2}'] = f'Итого (шт.):{counter_sum}'
        ws[f'K{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'K{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_communications (self):

        """
         заполенение таблицы комуникации

        """
        row = 9
        ws = self.wb['коммуникации']
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue

            list_comunications = v1.get('Воздушная коммуникация', {}).get('Вид коммуникации', []) + v1.get(
                'Подземная коммуникация', {}).get('Вид коммуникации', [])
            list_comunications_owner = v1.get('Воздушная коммуникация', {}).get('Собственник', []) + v1.get(
                'Подземная коммуникация', {}).get('Собственник', [])
            list_comunications.sort(key = lambda x: (x[-2][0], x[-2][1]))
            list_comunications_owner.sort(key = lambda x: (x[-2][0], x[-2][1]))

            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:L{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1

            for idx, value in enumerate(list_comunications):
                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = value[-2][0]
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-2][1]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[-1][0]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = value[-1][1]
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = value[0]
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'] = ''  # как найти расстояние от бровки?
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = ''  # как найти пересечение с осью?
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                ws[f'K{row}'] = ''
                ws[f'K{row}'].border = self.table_cells_border
                ws[f'K{row}'].alignment = self.table_cells_aligment
                ws[f'K{row}'].font = self.table_cells_font

                ws[f'L{row}'] = list_comunications_owner[idx][0] if idx < len(list_comunications_owner) else ''
                ws[f'L{row}'].border = self.table_cells_border
                ws[f'L{row}'].alignment = self.table_cells_aligment
                ws[f'L{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
        ws[f'L{row + 2}'] = f'Итого (шт.):{counter_sum}'
        ws[f'L{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'L{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_pipes (self):
        """
        заполенение таблицы трубы
        :return:
        """
        on_road = 0
        on_the_left = 0
        on_the_right = 0
        row = 8
        ws = self.wb['трубы']
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:N{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1

            for idx, value in enumerate(v1.get('Водопропускная труба', {}).get('Расположение', [])):
                if value[0].lower() == 'на дороге':
                    on_road += 1
                elif value[0].lower() == 'на примыкании слева':
                    on_the_left += 1
                elif value[0].lower() == 'на примыкании справа':
                    on_the_right += 1
                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = value[-2][0]
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-2][1]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = v1.get('Водопропускная труба', {}).get('Тип водотока', [])[idx][0] if idx < len(v1.get(
                    'Водопропускная труба', {}).get('Тип водотока', [])) else ''
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = v1.get('Водопропускная труба', {}).get('Материал', [])[idx][0] if idx < len(v1.get(
                    'Водопропускная труба', {}).get('Материал', [])) else ''
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = v1.get('Водопропускная труба', {}).get('Число очков', [])[idx][0] if idx < len(v1.get(
                    'Водопропускная труба', {}).get('Число очков', [])) else ''
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'] = v1.get('Водопропускная труба', {}).get('Диаметр (Ширина)', [])[idx][0] if idx < len(
                    v1.get('Водопропускная труба', {}).get('Диаметр (Ширина)', [])) else ''
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = v1.get('Водопропускная труба', {}).get('Высота', [])[idx][0] if idx < len(v1.get(
                    'Водопропускная труба', {}).get('Высота', [])) else ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = v1.get('Водопропускная труба', {}).get('Длина трубы по лотку', [])[idx][0] if idx < len(
                    v1.get('Водопропускная труба', {}).get('Длина трубы по лотку', [])) else ''
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                ws[f'K{row}'] = v1.get('Водопропускная труба', {}).get('Техническое состояние', [])[idx][
                    0] if idx < len(v1.get('Водопропускная труба', {}).get('Техническое состояние', [])) else ''
                ws[f'K{row}'].border = self.table_cells_border
                ws[f'K{row}'].alignment = self.table_cells_aligment
                ws[f'K{row}'].font = self.table_cells_font

                ws[f'L{row}'] = value[0]
                ws[f'L{row}'].border = self.table_cells_border
                ws[f'L{row}'].alignment = self.table_cells_aligment
                ws[f'L{row}'].font = self.table_cells_font

                ws[f'M{row}'] = ''
                ws[f'M{row}'].border = self.table_cells_border
                ws[f'M{row}'].alignment = self.table_cells_aligment
                ws[f'M{row}'].font = self.table_cells_font

                ws[f'N{row}'] = ''
                ws[f'N{row}'].border = self.table_cells_border
                ws[f'N{row}'].alignment = self.table_cells_aligment
                ws[f'N{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
        ws[f'K{row + 2}'] = 'Итого (шт.):'
        ws[f'K{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'K{row + 2}'].font = self.cells_font_result

        ws[f'L{row + 2}'] = f'На дороге {on_road}'
        ws[f'L{row + 2}'].alignment = self.cells_result
        ws[f'L{row + 2}'].font = self.cells_font_result

        ws[f'L{row + 3}'] = f'На примыкании слева {on_the_left}'
        ws[f'L{row + 3}'].alignment = self.cells_result
        ws[f'L{row + 3}'].font = self.cells_font_result

        ws[f'L{row + 4}'] = f'На примыкании справа {on_the_right}'
        ws[f'L{row + 4}'].alignment = self.cells_result
        ws[f'L{row + 4}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_bridge (self):
        """
        заполенение таблицы мосты
        :return:
        """
        row = 8
        ws = self.wb['мосты']
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:O{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1

            for idx, value in enumerate(v1.get('Мостовое сооружение', {}).get('Тип', [])):
                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = value[-2][0]
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-2][1]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[0]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                ws[f'F{row}'] = v1.get('Мостовое сооружение', {}).get('Пересекаемое препятствие', [])[idx][0] if \
                    (idx < len(v1.get('Мостовое сооружение', {}).get('Пересекаемое препятствие', []))) else ''
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = v1.get('Мостовое сооружение', {}).get('Полная длина моста', [])[idx][0] if (idx <
                                                                                                            len(v1.get(
                                                                                                                'Мостовое сооружение',
                                                                                                                {}).get(
                                                                                                                'Полная длина моста',
                                                                                                                []))) else ''
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                list_material = []
                list_tip_spans = []
                for i in range(1, 7):
                    material = v1.get('Мостовое сооружение', {}).get(f'Пролетн.стр.{i} Материал', [])[idx][0] if (idx <
                                                                                                                  len(v1.get(
                                                                                                                      'Мостовое сооружение',
                                                                                                                      {}).get(
                                                                                                                      f'Пролетн.стр.{i} Материал',
                                                                                                                      []))) else None
                    tip_spans = \
                        v1.get('Мостовое сооружение', {}).get(f'Пролетн.стр.{i} Тип пролетных строений', [])[idx][
                            0] if (idx < len(
                            v1.get('Мостовое сооружение', {}).get(f'Пролетн.стр.{i} Тип пролетных строений',
                                                                  []))) else None
                    if tip_spans:
                        list_tip_spans.append(tip_spans)
                    if material:
                        list_material.append(material)
                list_material = set(list_material)
                list_tip_spans = set(list_tip_spans)
                ws[f'H{row}'] = ','.join(list_material) if list_material else ''
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = ','.join(list_tip_spans) if list_tip_spans else ''
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = v1.get('Мостовое сооружение', {}).get('Габарит (высота)', [])[idx][0] if (idx <
                                                                                                          len(v1.get(
                                                                                                              'Мостовое сооружение',
                                                                                                              {}).get(
                                                                                                              'Габарит (высота)',
                                                                                                              []))) else ''
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                ws[f'K{row}'] = v1.get('Мостовое сооружение', {}).get('Нормативная нагрузка', [])[idx][0] if (idx <
                                                                                                              len(v1.get(
                                                                                                                  'Мостовое сооружение',
                                                                                                                  {}).get(
                                                                                                                  'Нормативная нагрузка',
                                                                                                                  []))) else ''
                ws[f'K{row}'].border = self.table_cells_border
                ws[f'K{row}'].alignment = self.table_cells_aligment
                ws[f'K{row}'].font = self.table_cells_font

                ws[f'L{row}'] = v1.get('Мостовое сооружение', {}).get('Техническое состояние', [])[idx][0] if (idx <
                                                                                                               len(v1.get(
                                                                                                                   'Мостовое сооружение',
                                                                                                                   {}).get(
                                                                                                                   'Техническое состояние',
                                                                                                                   []))) else ''
                ws[f'L{row}'].border = self.table_cells_border
                ws[f'L{row}'].alignment = self.table_cells_aligment
                ws[f'L{row}'].font = self.table_cells_font

                ws[f'M{row}'] = ''
                ws[f'M{row}'].border = self.table_cells_border
                ws[f'M{row}'].alignment = self.table_cells_aligment
                ws[f'M{row}'].font = self.table_cells_font

                ws[f'N{row}'] = ''
                ws[f'N{row}'].border = self.table_cells_border
                ws[f'N{row}'].alignment = self.table_cells_aligment
                ws[f'N{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
        ws[f'L{row + 2}'] = f'Итого (шт.):{counter_sum}'
        ws[f'L{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
        ws[f'L{row + 2}'].font = self.cells_font_result

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_turns (self):
        """
        заполенение таблицы повороты
        :return:
        """
        row = 10
        row_angle = 9
        ws = self.wb['повороты']
        counter_sum = 0
        for k1, v1 in self.data.items():

            counter = 0
            if isinstance(v1, str):
                continue

            list_radius = [j for i in v1.get('Кривая', {}).get('№ угла', []) for j in v1.get('Кривая', {}).get('R', [])
                           if i[-2] == j[-2] and i[-1] == j[-1]]

            ws[f'D{row_angle}'] = 'НТ'
            ws[f'D{row_angle}'].border = self.table_cells_border
            ws[f'D{row_angle}'].alignment = self.table_cells_aligment

            ws[f'E{row_angle}'] = v1.get('Ось дороги').get('Начало трассы')[0][-2][0]
            ws[f'E{row_angle}'].border = self.table_cells_border
            ws[f'E{row_angle}'].alignment = self.table_cells_aligment

            ws[f'F{row_angle}'] = v1.get('Ось дороги').get('Начало трассы')[0][-2][1]
            ws[f'F{row_angle}'].border = self.table_cells_border
            ws[f'F{row_angle}'].alignment = self.table_cells_aligment

            ws[f'G{row_angle}'].border = self.table_cells_border
            ws.merge_cells(f'G{row_angle}:P{row_angle}')
            ws[f'G{row_angle}'].alignment = self.table_cells_aligment
            if len(self.data) > 2:
                ws[f'G{row_angle}'] = k1

            for idx, value in enumerate(v1.get('Кривая', {}).get('Положение вершины', [])):
                if float(value[0]) == 0.0:

                    ws[f'Q{row_angle}'].border = self.table_cells_border
                    ws[f'Q{row_angle + 1}'].border = self.table_cells_border
                    ws.merge_cells(f'Q{row_angle}:Q{row_angle + 1}')
                    ws[f'Q{row_angle}'] = round(value[1])
                    ws[f'Q{row_angle}'].alignment = self.table_cells_aligment

                    ws[f'R{row_angle}'].border = self.table_cells_border
                    ws[f'R{row_angle + 1}'].border = self.table_cells_border
                    ws.merge_cells(f'R{row_angle}:R{row_angle + 1}')
                    if (idx == 0):
                        ws[f'R{row_angle}'] = f"=F{row}-F{row_angle}"
                        ic('ifНТ', row_angle, row)
                    elif (idx == len(v1.get('Кривая', {}).get('Положение вершины', []))):
                        ws[f'R{row_angle}'] = f"=F{row}-F{row_angle - 1}"
                        ic('elifНТ', row_angle, row)
                    else:
                        ws[f'R{row_angle}'] = f"=F{row_angle + 1}-F{row_angle - 1}"
                        ic('elseНТ', row_angle, row)

                    ws[f'R{row_angle}'].alignment = self.table_cells_aligment

                    ws[f'S{row_angle}'].border = self.table_cells_border
                    ws[f'S{row_angle + 1}'].border = self.table_cells_border
                    ws.merge_cells(f'S{row_angle}:S{row_angle + 1}')
                    angel = math.degrees(abs(float(v1.get('Кривая', {}).get('Угол поворота', [])[idx][0])))
                    gradus = int(angel // 1)
                    minutes = int(((angel - gradus) * 60) // 1)
                    ws[f'S{row_angle}'] = f'{gradus}°' + f'{minutes}\''
                    ws[f'S{row_angle}'].alignment = self.table_cells_aligment
                    # row_angle += 2
                    continue
                else:
                    ws[f'Q{row_angle}'].border = self.table_cells_border
                    ws[f'Q{row_angle + 1}'].border = self.table_cells_border
                    ws.merge_cells(f'Q{row_angle}:Q{row_angle + 1}')

                    ws[f'R{row_angle}'].border = self.table_cells_border
                    ws[f'R{row_angle + 1}'].border = self.table_cells_border
                    ws.merge_cells(f'R{row_angle}:R{row_angle + 1}')
                    ws[f'S{row_angle}'].border = self.table_cells_border
                    ws[f'S{row_angle + 1}'].border = self.table_cells_border
                    ws.merge_cells(f'S{row_angle}:S{row_angle + 1}')
                    row_angle += 2

                counter += 1
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'D{row}:D{row + 1}')
                ws[f'D{row}'] = v1.get('Кривая', {}).get('№ угла', [])[idx][0]
                ws[f'D{row}'].alignment = self.table_cells_aligment

                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'E{row}:E{row + 1}')
                ws[f'E{row}'] = 0
                ws[f'E{row}'].alignment = self.table_cells_aligment

                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'F{row}:F{row + 1}')
                ws[f'F{row}'] = round(float(value[0]))
                ws[f'F{row}'].alignment = self.table_cells_aligment

                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'G{row}:G{row + 1}')
                ws[f'G{row}'].alignment = self.table_cells_aligment

                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'H{row}:H{row + 1}')
                ws[f'H{row}'].alignment = self.table_cells_aligment

                if float(v1.get('Кривая', {}).get('Угол поворота', [])[idx][0]) < 0:
                    angel = math.degrees(abs(float(v1.get('Кривая', {}).get('Угол поворота', [])[idx][0])))
                    gradus = int(angel // 1)
                    minutes = int(((angel - gradus) * 60) // 1)
                    ws[f'G{row}'] = f'{gradus}°' + f'{minutes}\''
                else:
                    angel = math.degrees(abs(float(v1.get('Кривая', {}).get('Угол поворота', [])[idx][0])))
                    gradus = int(angel // 1)
                    minutes = int(((angel - gradus) * 60) // 1)
                    ws[f'H{row}'] = f'{gradus}°' + f'{minutes}\''

                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'I{row}:I{row + 1}')
                ws[f'I{row}'] = round(abs(float(list_radius[idx][0])))
                ws[f'I{row}'].alignment = self.table_cells_aligment

                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'J{row}:J{row + 1}')
                ws[f'J{row}'] = round(float(v1.get('Кривая', {}).get('Lk', [])[idx][0]))
                ws[f'J{row}'].alignment = self.table_cells_aligment

                ws[f'K{row}'].border = self.table_cells_border
                ws[f'K{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'K{row}:K{row + 1}')
                ws[f'K{row}'] = round(float(v1.get('Кривая', {}).get('T_1', [])[idx][0]))
                ws[f'K{row}'].alignment = self.table_cells_aligment

                ws[f'L{row}'].border = self.table_cells_border
                ws[f'L{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'L{row}:L{row + 1}')
                ws[f'L{row}'] = round(float(v1.get('Кривая', {}).get('Б', [])[idx][0]))
                ws[f'L{row}'].alignment = self.table_cells_aligment

                ws[f'M{row}'].border = self.table_cells_border
                ws[f'M{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'M{row}:M{row + 1}')
                ws[f'M{row}'] = value[-2][0]
                ws[f'M{row}'].alignment = self.table_cells_aligment

                ws[f'N{row}'].border = self.table_cells_border
                ws[f'N{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'N{row}:N{row + 1}')
                ws[f'N{row}'] = value[-2][1]
                ws[f'N{row}'].alignment = self.table_cells_aligment

                ws[f'O{row}'].border = self.table_cells_border
                ws[f'O{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'O{row}:O{row + 1}')
                ws[f'O{row}'] = value[-1][0]
                ws[f'O{row}'].alignment = self.table_cells_aligment

                ws[f'P{row}'].border = self.table_cells_border
                ws[f'P{row + 1}'].border = self.table_cells_border
                ws.merge_cells(f'P{row}:P{row + 1}')
                ws[f'P{row}'] = value[-1][1]
                ws[f'P{row}'].alignment = self.table_cells_aligment

                row += 2

            row_angle += 2
            ic('КТ', row_angle, row)
            ws[f'D{row}'] = 'КТ'
            ws[f'D{row}'].border = self.table_cells_border
            ws[f'D{row}'].alignment = self.table_cells_aligment

            ws[f'E{row}'] = v1.get('Ось дороги').get('Начало трассы')[0][-1][0]
            ws[f'E{row}'].border = self.table_cells_border
            ws[f'E{row}'].alignment = self.table_cells_aligment

            ws[f'F{row}'] = v1.get('Ось дороги').get('Начало трассы')[0][-1][1]
            ws[f'F{row}'].border = self.table_cells_border
            ws[f'F{row}'].alignment = self.table_cells_aligment

            ws[f'G{row}'].border = self.table_cells_border
            ws.merge_cells(f'G{row}:P{row}')
            if len(self.data) > 2:
                ws[f'G{row}'] = k1
            ws[f'G{row}'].alignment = self.table_cells_aligment

            row += 2
            counter_sum += counter

        if counter_sum == 0:
            ws.sheet_state = 'hidden'

    def write_gazon (self):
        """
        заполнение ведомости газон
        :return:
        """

        row = 8
        ws = self.wb['газон']
        counter_sum = 0
        for k1, v1 in self.data.items():
            counter = 0
            if isinstance(v1, str):
                continue
            if len(self.data) > 2:
                ws.merge_cells(f'B{row}:J{row}')
                ws[f'B{row}'] = k1
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font
                row += 1

            for value in v1.get('Газон', {}).get('Вид газона', []):
                counter += 1
                ws[f'B{row}'] = counter
                ws[f'B{row}'].border = self.table_cells_border
                ws[f'B{row}'].alignment = self.table_cells_aligment
                ws[f'B{row}'].font = self.table_cells_font

                ws[f'C{row}'] = 'чет.' if value[6] > 0 else 'нечет.'
                ws[f'C{row}'].border = self.table_cells_border
                ws[f'C{row}'].alignment = self.table_cells_aligment
                ws[f'C{row}'].font = self.table_cells_font

                ws[f'D{row}'] = value[-4]
                ws[f'D{row}'].border = self.table_cells_border
                ws[f'D{row}'].alignment = self.table_cells_aligment
                ws[f'D{row}'].font = self.table_cells_font

                ws[f'E{row}'] = value[-3]
                ws[f'E{row}'].border = self.table_cells_border
                ws[f'E{row}'].alignment = self.table_cells_aligment
                ws[f'E{row}'].font = self.table_cells_font

                width = round(value[2] / (value[-3] - value[-4]), 2)
                ws[f'F{row}'] = width
                ws[f'F{row}'].border = self.table_cells_border
                ws[f'F{row}'].alignment = self.table_cells_aligment
                ws[f'F{row}'].font = self.table_cells_font

                ws[f'G{row}'] = width
                ws[f'G{row}'].border = self.table_cells_border
                ws[f'G{row}'].alignment = self.table_cells_aligment
                ws[f'G{row}'].font = self.table_cells_font

                ws[f'H{row}'] = value[0]
                ws[f'H{row}'].border = self.table_cells_border
                ws[f'H{row}'].alignment = self.table_cells_aligment
                ws[f'H{row}'].font = self.table_cells_font

                ws[f'I{row}'] = value[2]
                ws[f'I{row}'].border = self.table_cells_border
                ws[f'I{row}'].alignment = self.table_cells_aligment
                ws[f'I{row}'].font = self.table_cells_font

                ws[f'J{row}'] = ''
                ws[f'J{row}'].border = self.table_cells_border
                ws[f'J{row}'].alignment = self.table_cells_aligment
                ws[f'J{row}'].font = self.table_cells_font

                row += 1

            counter_sum += counter
            ws[f'J{row + 2}'] = f'Итого (шт.):{counter_sum}'
            ws[f'J{row + 2}'].alignment = self.cells_result  # выравнивание по правому краю
            ws[f'J{row + 2}'].font = self.cells_font_result

            if counter_sum == 0:
                ws.sheet_state = 'hidden'


class WriterApplicationNotCityTP(WriterApplication):
    def __init__ (self, data: dict = None):
        super().__init__(data)
        # path_template_excel_application = ''
        self.wb = load_workbook(path_template_excel_application, keep_vba = True)


def convert_visio2svg (path_dir, name, ):
    """ Конвертируем визио в svg"""
    visio = win32com.client.gencache.EnsureDispatch("Visio.Application")
    doc = visio.Documents.Open(rf"{path_dir}\{name}.vsd")
    print(rf"{path_dir}\{name}.vsd")

    for page in doc.Pages:
        print(rf'{path_dir}\{page.Name}.pdf')
        page.Export(rf"{path_dir}\{page}.png")
    visio.Quit()


def new_titel (name_road):
    import win32com.client

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    wb = excel.Workbooks.Open(r"C:\Users\sibregion\Desktop\test\report\тест_рамок_пдф\ТИТУТЛЬНИК_ИЖЕВСК.xlsx")
    sheet = wb.Worksheets('Переплет')
    sheet.Cells(22, 2).Value = name_road
    wb.SaveCopyAs(fr"C:\Users\sibregion\Desktop\test\report\тест_рамок_пдф\ТИТУТЛЬНИК{name_road}.xlsx")

    excel.Application.Quit()


def main ():
    conn = db.Query('OMSK_REGION_2024')  # 'IZHEVSK_CITY_2023'
    list_roads = conn.set_road_name()
    data = conn.get_dad_datas(list_roads[5])  # 'М14 Е105 - до с.Вишневое'
    # print(data)
    conn.close_db()
    diсt_inter = {'year': 2024,
                  'tip_passport': 'city',
                  'history_match': '',
                  'area_conditioins': ''}
    # reports = WriterExcelTP(data = data, path = r'C:\Users\sibregion\Desktop\test\ReportGenerator-new_version',
    #                         data_interface = diсt_inter)
    # apps = WriterApplicationCityTP(data = data, path = r'C:\Users\sibregion\Desktop\test\ReportGenerator-new_version',
    #                         data_interface = diсt_inter)
    dad = WriterExcelDAD(data = data, path = r'C:\Users\sibregion\Desktop\test\ReportGenerator-new_version',
                         data_interface = diсt_inter)

    # apps = WriterApplicationCityTP(data = data,
    #                                path = r'C:\Users\sibregion\Desktop\test\report\тест_рамок_пдф\Приложения')
    # apps.save_file()

    # for name in conn.get_list_roads():
    #     print(name)
    #
    #     #time.sleep(10)
    #     try:
    #
    #         data = conn.get_tp_datas(name)
    #         apps = WriterApplicationCityTP(data = data, path = r'C:\Приложения',data_interface = dcit_inter)
    #         apps.save_file()
    #
    #
    #     except Exception as e:
    #         list_errors.append(name)
    #         print('error',name)
    #         print(e)
    #
    #         continue
    #     finally:
    #         print(list_errors)
    #         print(len(list_errors))
    # data = conn.get_tp_datas('ул. Масленникова')

    # report_application = WriterApplication(data)
    # report = WriterExcelTP(data,path =r'C:\Users\sibregion\Desktop\test\report')
    # report.save_file()
    # apps = WriterApplicationCityTP(data=data,path = r'\\Sibregion-srv2\отчеты\1.ТП\2023\г. Ижевск\1.Приложения')
    # apps.save_file()


if __name__ == "__main__":
    import time

    start = time.time()  # точка отсчета времени
    main()
    end = time.time() - start  # собственно время работы программы
    print(f"{round(end, 1)} секунд")  # вывод времени
